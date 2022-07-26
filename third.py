import itertools
import threading
import time
import warnings

import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
import os
import mysql.connector
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidgetItem, QAbstractItemView, QScrollBar, QMessageBox
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5 import QtWidgets
warnings.filterwarnings("ignore", category=DeprecationWarning)
from Instagram_FINAL.highlights_delete import highlights_delete
from Instagram_FINAL.seq_initial_login import seq_initial_login
from Instagram_FINAL.analytics import analytics
from Instagram_FINAL.first_login import first_login
from Instagram_FINAL.reel_posting import reel_posting
from Instagram_FINAL.reset_proxy import reset_proxy
from Instagram_FINAL.save_excel import save_excel
from Instagram_FINAL.threaded_dm import dm
from Instagram_FINAL.threaded_follow_only import threaded_follow_only
from Instagram_FINAL.threaded_mentions import mentions
from Instagram_FINAL.threaded_posts import posts
from Instagram_FINAL.threaded_relogin import threaded_relogin
from Instagram_FINAL.threaded_story import threaded_story
from Instagram_FINAL.warmup import warmup

from PyQt5 import QtCore, QtGui, QtWidgets


class Ui_MWin(object):
    def setupUi(self, MWin):
        MWin.setObjectName("MWin")
        MWin.resize(1819, 1148)
        MWin.setMinimumSize(QtCore.QSize(720, 400))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        MWin.setFont(font)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/images/music-box.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MWin.setWindowIcon(icon)
        MWin.setIconSize(QtCore.QSize(40, 40))
        self.centralWidget = QtWidgets.QWidget(MWin)
        self.centralWidget.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.centralWidget.setObjectName("centralWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralWidget)
        self.verticalLayout.setContentsMargins(5, 5, 0, 5)
        self.verticalLayout.setSpacing(5)
        self.verticalLayout.setObjectName("verticalLayout")
        self.Rough = QtWidgets.QTabWidget(self.centralWidget)
        self.Rough.setObjectName("Rough")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.frame = QtWidgets.QFrame(self.tab)
        self.frame.setGeometry(QtCore.QRect(60, 20, 941, 381))
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.layoutWidget = QtWidgets.QWidget(self.frame)
        self.layoutWidget.setGeometry(QtCore.QRect(30, 70, 191, 131))
        self.layoutWidget.setObjectName("layoutWidget")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.layoutWidget)
        self.gridLayout_3.setContentsMargins(11, 11, 11, 11)
        self.gridLayout_3.setSpacing(6)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.label_6 = QtWidgets.QLabel(self.layoutWidget)
        self.label_6.setObjectName("label_6")
        self.gridLayout_3.addWidget(self.label_6, 1, 0, 1, 1)
        self.label_28 = QtWidgets.QLabel(self.layoutWidget)
        self.label_28.setObjectName("label_28")
        self.gridLayout_3.addWidget(self.label_28, 2, 0, 1, 1)
        self.label_5 = QtWidgets.QLabel(self.layoutWidget)
        self.label_5.setObjectName("label_5")
        self.gridLayout_3.addWidget(self.label_5, 0, 0, 1, 1)
        self.label_7 = QtWidgets.QLabel(self.layoutWidget)
        self.label_7.setObjectName("label_7")
        self.gridLayout_3.addWidget(self.label_7, 3, 0, 1, 1)
        self.label_32 = QtWidgets.QLabel(self.layoutWidget)
        self.label_32.setObjectName("label_32")
        self.gridLayout_3.addWidget(self.label_32, 4, 0, 1, 1)
        self.layoutWidget1 = QtWidgets.QWidget(self.frame)
        self.layoutWidget1.setGeometry(QtCore.QRect(220, 70, 81, 136))
        self.layoutWidget1.setObjectName("layoutWidget1")
        self.gridLayout = QtWidgets.QGridLayout(self.layoutWidget1)
        self.gridLayout.setContentsMargins(11, 11, 11, 11)
        self.gridLayout.setSpacing(6)
        self.gridLayout.setObjectName("gridLayout")
        self.notallocated_acc_summary = QtWidgets.QLabel(self.layoutWidget1)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.notallocated_acc_summary.setFont(font)
        self.notallocated_acc_summary.setText("")
        self.notallocated_acc_summary.setObjectName("notallocated_acc_summary")
        self.gridLayout.addWidget(self.notallocated_acc_summary, 2, 0, 1, 1)
        self.active_acc_summary = QtWidgets.QLabel(self.layoutWidget1)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.active_acc_summary.setFont(font)
        self.active_acc_summary.setText("")
        self.active_acc_summary.setObjectName("active_acc_summary")
        self.gridLayout.addWidget(self.active_acc_summary, 1, 0, 1, 1)
        self.challenge_acc_summary = QtWidgets.QLabel(self.layoutWidget1)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.challenge_acc_summary.setFont(font)
        self.challenge_acc_summary.setText("")
        self.challenge_acc_summary.setObjectName("challenge_acc_summary")
        self.gridLayout.addWidget(self.challenge_acc_summary, 3, 0, 1, 1)
        self.added_acc_summary = QtWidgets.QLabel(self.layoutWidget1)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.added_acc_summary.setFont(font)
        self.added_acc_summary.setText("")
        self.added_acc_summary.setObjectName("added_acc_summary")
        self.gridLayout.addWidget(self.added_acc_summary, 0, 0, 1, 1)
        self.totalcategory_acc_summary = QtWidgets.QLabel(self.layoutWidget1)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.totalcategory_acc_summary.setFont(font)
        self.totalcategory_acc_summary.setText("")
        self.totalcategory_acc_summary.setObjectName("totalcategory_acc_summary")
        self.gridLayout.addWidget(self.totalcategory_acc_summary, 4, 0, 1, 1)
        self.label_3 = QtWidgets.QLabel(self.frame)
        self.label_3.setGeometry(QtCore.QRect(110, 20, 161, 20))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.add_acc_path_button = QtWidgets.QToolButton(self.frame)
        self.add_acc_path_button.setGeometry(QtCore.QRect(30, 230, 31, 31))
        self.add_acc_path_button.setObjectName("add_acc_path_button")
        self.add_acc_path = QtWidgets.QLabel(self.frame)
        self.add_acc_path.setGeometry(QtCore.QRect(70, 240, 711, 16))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(7)
        self.add_acc_path.setFont(font)
        self.add_acc_path.setText("")
        self.add_acc_path.setObjectName("add_acc_path")
        self.label_22 = QtWidgets.QLabel(self.frame)
        self.label_22.setGeometry(QtCore.QRect(30, 270, 115, 21))
        self.label_22.setObjectName("label_22")
        self.addgroup_acc_combobox = QtWidgets.QComboBox(self.frame)
        self.addgroup_acc_combobox.setGeometry(QtCore.QRect(150, 270, 161, 26))
        self.addgroup_acc_combobox.setObjectName("addgroup_acc_combobox")
        self.addgroup_acc_combobox.addItem("")
        self.frame_2 = QtWidgets.QFrame(self.tab)
        self.frame_2.setGeometry(QtCore.QRect(30, 420, 961, 491))
        self.frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.label_4 = QtWidgets.QLabel(self.frame_2)
        self.label_4.setGeometry(QtCore.QRect(120, 20, 131, 20))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.add_proxy_path_button = QtWidgets.QToolButton(self.frame_2)
        self.add_proxy_path_button.setGeometry(QtCore.QRect(40, 190, 31, 31))
        self.add_proxy_path_button.setObjectName("add_proxy_path_button")
        self.layoutWidget2 = QtWidgets.QWidget(self.frame_2)
        self.layoutWidget2.setGeometry(QtCore.QRect(230, 70, 71, 100))
        self.layoutWidget2.setObjectName("layoutWidget2")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.layoutWidget2)
        self.gridLayout_4.setContentsMargins(11, 11, 11, 11)
        self.gridLayout_4.setSpacing(6)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.added_proxy_summary = QtWidgets.QLabel(self.layoutWidget2)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.added_proxy_summary.setFont(font)
        self.added_proxy_summary.setText("")
        self.added_proxy_summary.setObjectName("added_proxy_summary")
        self.gridLayout_4.addWidget(self.added_proxy_summary, 0, 0, 1, 1)
        self.active_proxy_summary = QtWidgets.QLabel(self.layoutWidget2)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.active_proxy_summary.setFont(font)
        self.active_proxy_summary.setText("")
        self.active_proxy_summary.setObjectName("active_proxy_summary")
        self.gridLayout_4.addWidget(self.active_proxy_summary, 1, 0, 1, 1)
        self.assigned_proxy_summary = QtWidgets.QLabel(self.layoutWidget2)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.assigned_proxy_summary.setFont(font)
        self.assigned_proxy_summary.setText("")
        self.assigned_proxy_summary.setObjectName("assigned_proxy_summary")
        self.gridLayout_4.addWidget(self.assigned_proxy_summary, 2, 0, 1, 1)
        self.notassigned_proxy_summary = QtWidgets.QLabel(self.layoutWidget2)
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.notassigned_proxy_summary.setFont(font)
        self.notassigned_proxy_summary.setText("")
        self.notassigned_proxy_summary.setObjectName("notassigned_proxy_summary")
        self.gridLayout_4.addWidget(self.notassigned_proxy_summary, 3, 0, 1, 1)
        self.layoutWidget3 = QtWidgets.QWidget(self.frame_2)
        self.layoutWidget3.setGeometry(QtCore.QRect(40, 70, 191, 100))
        self.layoutWidget3.setObjectName("layoutWidget3")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.layoutWidget3)
        self.gridLayout_2.setContentsMargins(11, 11, 11, 11)
        self.gridLayout_2.setSpacing(6)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.label_9 = QtWidgets.QLabel(self.layoutWidget3)
        self.label_9.setObjectName("label_9")
        self.gridLayout_2.addWidget(self.label_9, 0, 0, 1, 1)
        self.label_10 = QtWidgets.QLabel(self.layoutWidget3)
        self.label_10.setObjectName("label_10")
        self.gridLayout_2.addWidget(self.label_10, 1, 0, 1, 1)
        self.label_8 = QtWidgets.QLabel(self.layoutWidget3)
        self.label_8.setObjectName("label_8")
        self.gridLayout_2.addWidget(self.label_8, 2, 0, 1, 1)
        self.label_11 = QtWidgets.QLabel(self.layoutWidget3)
        self.label_11.setObjectName("label_11")
        self.gridLayout_2.addWidget(self.label_11, 3, 0, 1, 1)
        self.add_proxy_path = QtWidgets.QLabel(self.frame_2)
        self.add_proxy_path.setGeometry(QtCore.QRect(80, 200, 551, 16))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(7)
        self.add_proxy_path.setFont(font)
        self.add_proxy_path.setText("")
        self.add_proxy_path.setObjectName("add_proxy_path")
        self.listWidget = QtWidgets.QListWidget(self.frame_2)
        self.listWidget.setGeometry(QtCore.QRect(40, 260, 431, 192))
        self.listWidget.setObjectName("listWidget")
        self.add_proxy_path_2 = QtWidgets.QLabel(self.frame_2)
        self.add_proxy_path_2.setGeometry(QtCore.QRect(40, 230, 91, 16))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(9)
        self.add_proxy_path_2.setFont(font)
        self.add_proxy_path_2.setObjectName("add_proxy_path_2")
        self.account_proxy_dial = QtWidgets.QSpinBox(self.frame_2)
        self.account_proxy_dial.setGeometry(QtCore.QRect(470, 100, 91, 22))
        self.account_proxy_dial.setObjectName("account_proxy_dial")
        self.label_127 = QtWidgets.QLabel(self.frame_2)
        self.label_127.setGeometry(QtCore.QRect(310, 100, 151, 21))
        self.label_127.setObjectName("label_127")
        self.label_128 = QtWidgets.QLabel(self.frame_2)
        self.label_128.setGeometry(QtCore.QRect(310, 70, 151, 21))
        self.label_128.setObjectName("label_128")
        self.account_proxy_dial_4 = QtWidgets.QSpinBox(self.frame_2)
        self.account_proxy_dial_4.setGeometry(QtCore.QRect(470, 70, 91, 22))
        self.account_proxy_dial_4.setObjectName("account_proxy_dial_4")
        self.pushButton_8 = QtWidgets.QPushButton(self.frame_2)
        self.pushButton_8.setGeometry(QtCore.QRect(470, 140, 91, 31))
        self.pushButton_8.setObjectName("pushButton_8")
        self.Rough.addTab(self.tab, "")
        self.tab2 = QtWidgets.QWidget()
        self.tab2.setObjectName("tab2")
        self.dtable = QtWidgets.QTableWidget(self.tab2)
        self.dtable.setGeometry(QtCore.QRect(0, 70, 1801, 921))
        self.dtable.setAlternatingRowColors(False)
        self.dtable.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.dtable.setGridStyle(QtCore.Qt.SolidLine)
        self.dtable.setObjectName("dtable")
        self.dtable.setColumnCount(0)
        self.dtable.setRowCount(0)
        self.dtable.horizontalHeader().setVisible(False)
        self.dtable.horizontalHeader().setCascadingSectionResizes(True)
        self.dtable.horizontalHeader().setHighlightSections(False)
        self.dtable.horizontalHeader().setMinimumSectionSize(37)
        self.dtable.horizontalHeader().setSortIndicatorShown(False)
        self.dtable.horizontalHeader().setStretchLastSection(True)
        self.dtable.verticalHeader().setVisible(False)
        self.dtable.verticalHeader().setCascadingSectionResizes(False)
        self.dtable.verticalHeader().setSortIndicatorShown(True)
        self.label = QtWidgets.QLabel(self.tab2)
        self.label.setGeometry(QtCore.QRect(10, 10, 101, 31))
        self.label.setObjectName("label")
        self.acc_selected_label = QtWidgets.QLabel(self.tab2)
        self.acc_selected_label.setGeometry(QtCore.QRect(120, 10, 71, 31))
        self.acc_selected_label.setObjectName("acc_selected_label")
        self.start_button = QtWidgets.QToolButton(self.tab2)
        self.start_button.setGeometry(QtCore.QRect(1640, 40, 71, 26))
        self.start_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.start_button.setObjectName("start_button")
        self.delete_acc_button = QtWidgets.QToolButton(self.tab2)
        self.delete_acc_button.setGeometry(QtCore.QRect(1170, 40, 81, 26))
        self.delete_acc_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.delete_acc_button.setObjectName("delete_acc_button")
        self.data_reload_button = QtWidgets.QToolButton(self.tab2)
        self.data_reload_button.setGeometry(QtCore.QRect(1080, 40, 66, 26))
        self.data_reload_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.data_reload_button.setObjectName("data_reload_button")
        self.group_acc_combobox = QtWidgets.QComboBox(self.tab2)
        self.group_acc_combobox.setGeometry(QtCore.QRect(540, 10, 161, 26))
        self.group_acc_combobox.setObjectName("group_acc_combobox")
        self.group_acc_combobox.addItem("")
        self.label_20 = QtWidgets.QLabel(self.tab2)
        self.label_20.setGeometry(QtCore.QRect(410, 10, 115, 21))
        self.label_20.setObjectName("label_20")
        self.activity_combobox = QtWidgets.QComboBox(self.tab2)
        self.activity_combobox.setGeometry(QtCore.QRect(1540, 10, 171, 21))
        self.activity_combobox.setObjectName("activity_combobox")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.activity_combobox.addItem("")
        self.label_21 = QtWidgets.QLabel(self.tab2)
        self.label_21.setGeometry(QtCore.QRect(1470, 10, 61, 20))
        self.label_21.setObjectName("label_21")
        self.label_34 = QtWidgets.QLabel(self.tab2)
        self.label_34.setGeometry(QtCore.QRect(710, 0, 115, 21))
        self.label_34.setObjectName("label_34")
        self.changegroup_acc_combobox = QtWidgets.QComboBox(self.tab2)
        self.changegroup_acc_combobox.setGeometry(QtCore.QRect(830, 0, 151, 26))
        self.changegroup_acc_combobox.setObjectName("changegroup_acc_combobox")
        self.changegroup_acc_combobox.addItem("")
        self.changegroup_acc_combobox.addItem("")
        self.groupchange_btn = QtWidgets.QToolButton(self.tab2)
        self.groupchange_btn.setGeometry(QtCore.QRect(1000, 10, 66, 41))
        self.groupchange_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.groupchange_btn.setObjectName("groupchange_btn")
        self.accounts_allocated_label = QtWidgets.QLabel(self.tab2)
        self.accounts_allocated_label.setGeometry(QtCore.QRect(570, 40, 121, 21))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.accounts_allocated_label.setFont(font)
        self.accounts_allocated_label.setObjectName("accounts_allocated_label")
        self.label_44 = QtWidgets.QLabel(self.tab2)
        self.label_44.setGeometry(QtCore.QRect(410, 40, 151, 21))
        self.label_44.setObjectName("label_44")
        self.layoutWidget4 = QtWidgets.QWidget(self.tab2)
        self.layoutWidget4.setGeometry(QtCore.QRect(200, 10, 183, 56))
        self.layoutWidget4.setObjectName("layoutWidget4")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.layoutWidget4)
        self.verticalLayout_2.setContentsMargins(11, 11, 11, 11)
        self.verticalLayout_2.setSpacing(6)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.selectall_acc_checkbox = QtWidgets.QCheckBox(self.layoutWidget4)
        self.selectall_acc_checkbox.setObjectName("selectall_acc_checkbox")
        self.verticalLayout_2.addWidget(self.selectall_acc_checkbox)
        self.deselectall_acc_checkbox = QtWidgets.QCheckBox(self.layoutWidget4)
        self.deselectall_acc_checkbox.setObjectName("deselectall_acc_checkbox")
        self.verticalLayout_2.addWidget(self.deselectall_acc_checkbox)
        self.label_38 = QtWidgets.QLabel(self.tab2)
        self.label_38.setGeometry(QtCore.QRect(710, 40, 115, 21))
        self.label_38.setObjectName("label_38")
        self.groupchange_textedit_textbox_2 = QtWidgets.QTextEdit(self.tab2)
        self.groupchange_textedit_textbox_2.setGeometry(QtCore.QRect(830, 30, 151, 31))
        self.groupchange_textedit_textbox_2.setObjectName("groupchange_textedit_textbox_2")
        self.account_proxy_dial_6 = QtWidgets.QSpinBox(self.tab2)
        self.account_proxy_dial_6.setGeometry(QtCore.QRect(1350, 10, 91, 22))
        self.account_proxy_dial_6.setObjectName("account_proxy_dial_6")
        self.account_proxy_dial_7 = QtWidgets.QSpinBox(self.tab2)
        self.account_proxy_dial_7.setGeometry(QtCore.QRect(1350, 40, 91, 22))
        self.account_proxy_dial_7.setObjectName("account_proxy_dial_7")
        self.label_129 = QtWidgets.QLabel(self.tab2)
        self.label_129.setGeometry(QtCore.QRect(1270, 10, 71, 21))
        self.label_129.setObjectName("label_129")
        self.label_130 = QtWidgets.QLabel(self.tab2)
        self.label_130.setGeometry(QtCore.QRect(1270, 40, 71, 21))
        self.label_130.setObjectName("label_130")
        self.selectall_acc_checkbox_2 = QtWidgets.QCheckBox(self.tab2)
        self.selectall_acc_checkbox_2.setGeometry(QtCore.QRect(1080, 10, 181, 24))
        self.selectall_acc_checkbox_2.setObjectName("selectall_acc_checkbox_2")
        self.Rough.addTab(self.tab2, "")
        self.tab_3 = QtWidgets.QWidget()
        self.tab_3.setObjectName("tab_3")
        self.account_group_frame = QtWidgets.QFrame(self.tab_3)
        self.account_group_frame.setGeometry(QtCore.QRect(630, 40, 371, 271))
        self.account_group_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.account_group_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.account_group_frame.setObjectName("account_group_frame")
        self.label_35 = QtWidgets.QLabel(self.account_group_frame)
        self.label_35.setGeometry(QtCore.QRect(90, 20, 151, 16))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.label_35.setFont(font)
        self.label_35.setObjectName("label_35")
        self.group_acc_combobox_2 = QtWidgets.QComboBox(self.account_group_frame)
        self.group_acc_combobox_2.setGeometry(QtCore.QRect(20, 51, 271, 31))
        self.group_acc_combobox_2.setObjectName("group_acc_combobox_2")
        self.groupchange_textedit_textbox = QtWidgets.QTextEdit(self.account_group_frame)
        self.groupchange_textedit_textbox.setGeometry(QtCore.QRect(120, 150, 171, 31))
        self.groupchange_textedit_textbox.setObjectName("groupchange_textedit_textbox")
        self.label_36 = QtWidgets.QLabel(self.account_group_frame)
        self.label_36.setGeometry(QtCore.QRect(20, 150, 91, 31))
        self.label_36.setObjectName("label_36")
        self.label_37 = QtWidgets.QLabel(self.account_group_frame)
        self.label_37.setGeometry(QtCore.QRect(20, 110, 151, 21))
        self.label_37.setObjectName("label_37")
        self.accounts_allocated_label_2 = QtWidgets.QLabel(self.account_group_frame)
        self.accounts_allocated_label_2.setGeometry(QtCore.QRect(180, 110, 71, 21))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.accounts_allocated_label_2.setFont(font)
        self.accounts_allocated_label_2.setObjectName("accounts_allocated_label_2")
        self.mentions_settings_frame = QtWidgets.QFrame(self.tab_3)
        self.mentions_settings_frame.setGeometry(QtCore.QRect(1060, 420, 381, 351))
        self.mentions_settings_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.mentions_settings_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.mentions_settings_frame.setObjectName("mentions_settings_frame")
        self.targetuser_mention_button = QtWidgets.QPushButton(self.mentions_settings_frame)
        self.targetuser_mention_button.setGeometry(QtCore.QRect(40, 100, 41, 28))
        self.targetuser_mention_button.setObjectName("targetuser_mention_button")
        self.label_91 = QtWidgets.QLabel(self.mentions_settings_frame)
        self.label_91.setGeometry(QtCore.QRect(40, 70, 181, 31))
        self.label_91.setObjectName("label_91")
        self.label_92 = QtWidgets.QLabel(self.mentions_settings_frame)
        self.label_92.setGeometry(QtCore.QRect(40, 30, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_92.setFont(font)
        self.label_92.setObjectName("label_92")
        self.targetuser_mention_path = QtWidgets.QLabel(self.mentions_settings_frame)
        self.targetuser_mention_path.setGeometry(QtCore.QRect(90, 105, 421, 21))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(7)
        self.targetuser_mention_path.setFont(font)
        self.targetuser_mention_path.setText("")
        self.targetuser_mention_path.setObjectName("targetuser_mention_path")
        self.comment_mentions_spinBox = QtWidgets.QSpinBox(self.mentions_settings_frame)
        self.comment_mentions_spinBox.setGeometry(QtCore.QRect(300, 210, 42, 22))
        self.comment_mentions_spinBox.setObjectName("comment_mentions_spinBox")
        self.label_96 = QtWidgets.QLabel(self.mentions_settings_frame)
        self.label_96.setGeometry(QtCore.QRect(40, 210, 181, 21))
        self.label_96.setObjectName("label_96")
        self.label_98 = QtWidgets.QLabel(self.mentions_settings_frame)
        self.label_98.setGeometry(QtCore.QRect(40, 240, 261, 21))
        self.label_98.setObjectName("label_98")
        self.frequency_mentions_spinBox = QtWidgets.QSpinBox(self.mentions_settings_frame)
        self.frequency_mentions_spinBox.setGeometry(QtCore.QRect(300, 240, 42, 22))
        self.frequency_mentions_spinBox.setObjectName("frequency_mentions_spinBox")
        self.label_94 = QtWidgets.QLabel(self.mentions_settings_frame)
        self.label_94.setGeometry(QtCore.QRect(40, 140, 191, 21))
        self.label_94.setObjectName("label_94")
        self.message_dm_textbox_2 = QtWidgets.QTextEdit(self.mentions_settings_frame)
        self.message_dm_textbox_2.setGeometry(QtCore.QRect(40, 170, 301, 31))
        self.message_dm_textbox_2.setObjectName("message_dm_textbox_2")
        self.pushButton_2 = QtWidgets.QPushButton(self.mentions_settings_frame)
        self.pushButton_2.setGeometry(QtCore.QRect(270, 290, 75, 23))
        self.pushButton_2.setObjectName("pushButton_2")
        self.dm_settings_frame = QtWidgets.QFrame(self.tab_3)
        self.dm_settings_frame.setGeometry(QtCore.QRect(630, 420, 381, 371))
        self.dm_settings_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.dm_settings_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.dm_settings_frame.setObjectName("dm_settings_frame")
        self.targetuser_dm_button = QtWidgets.QPushButton(self.dm_settings_frame)
        self.targetuser_dm_button.setGeometry(QtCore.QRect(40, 100, 41, 28))
        self.targetuser_dm_button.setObjectName("targetuser_dm_button")
        self.label_95 = QtWidgets.QLabel(self.dm_settings_frame)
        self.label_95.setGeometry(QtCore.QRect(40, 70, 181, 31))
        self.label_95.setObjectName("label_95")
        self.label_99 = QtWidgets.QLabel(self.dm_settings_frame)
        self.label_99.setGeometry(QtCore.QRect(40, 30, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_99.setFont(font)
        self.label_99.setObjectName("label_99")
        self.targetuser_dm_path = QtWidgets.QLabel(self.dm_settings_frame)
        self.targetuser_dm_path.setGeometry(QtCore.QRect(90, 105, 431, 21))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(7)
        self.targetuser_dm_path.setFont(font)
        self.targetuser_dm_path.setText("")
        self.targetuser_dm_path.setObjectName("targetuser_dm_path")
        self.label_102 = QtWidgets.QLabel(self.dm_settings_frame)
        self.label_102.setGeometry(QtCore.QRect(40, 280, 261, 21))
        self.label_102.setObjectName("label_102")
        self.frequency_dm_account = QtWidgets.QSpinBox(self.dm_settings_frame)
        self.frequency_dm_account.setGeometry(QtCore.QRect(300, 280, 42, 22))
        self.frequency_dm_account.setObjectName("frequency_dm_account")
        self.label_104 = QtWidgets.QLabel(self.dm_settings_frame)
        self.label_104.setGeometry(QtCore.QRect(40, 140, 191, 21))
        self.label_104.setObjectName("label_104")
        self.message_dm_textbox = QtWidgets.QTextEdit(self.dm_settings_frame)
        self.message_dm_textbox.setGeometry(QtCore.QRect(40, 170, 321, 101))
        self.message_dm_textbox.setObjectName("message_dm_textbox")
        self.pushButton_3 = QtWidgets.QPushButton(self.dm_settings_frame)
        self.pushButton_3.setGeometry(QtCore.QRect(290, 320, 75, 23))
        self.pushButton_3.setObjectName("pushButton_3")
        self.dm_settings_frame_2 = QtWidgets.QFrame(self.tab_3)
        self.dm_settings_frame_2.setGeometry(QtCore.QRect(1040, 40, 381, 241))
        self.dm_settings_frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.dm_settings_frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.dm_settings_frame_2.setObjectName("dm_settings_frame_2")
        self.targetuser_dm_button_2 = QtWidgets.QPushButton(self.dm_settings_frame_2)
        self.targetuser_dm_button_2.setGeometry(QtCore.QRect(40, 100, 41, 28))
        self.targetuser_dm_button_2.setObjectName("targetuser_dm_button_2")
        self.label_97 = QtWidgets.QLabel(self.dm_settings_frame_2)
        self.label_97.setGeometry(QtCore.QRect(40, 70, 181, 31))
        self.label_97.setObjectName("label_97")
        self.label_100 = QtWidgets.QLabel(self.dm_settings_frame_2)
        self.label_100.setGeometry(QtCore.QRect(40, 30, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_100.setFont(font)
        self.label_100.setObjectName("label_100")
        self.targetuser_dm_path_2 = QtWidgets.QLabel(self.dm_settings_frame_2)
        self.targetuser_dm_path_2.setGeometry(QtCore.QRect(90, 105, 391, 21))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(7)
        self.targetuser_dm_path_2.setFont(font)
        self.targetuser_dm_path_2.setText("")
        self.targetuser_dm_path_2.setObjectName("targetuser_dm_path_2")
        self.label_103 = QtWidgets.QLabel(self.dm_settings_frame_2)
        self.label_103.setGeometry(QtCore.QRect(30, 140, 261, 21))
        self.label_103.setObjectName("label_103")
        self.frequency_dm_account_2 = QtWidgets.QSpinBox(self.dm_settings_frame_2)
        self.frequency_dm_account_2.setGeometry(QtCore.QRect(280, 140, 71, 22))
        self.frequency_dm_account_2.setObjectName("frequency_dm_account_2")
        self.label_114 = QtWidgets.QLabel(self.dm_settings_frame_2)
        self.label_114.setGeometry(QtCore.QRect(30, 170, 261, 21))
        self.label_114.setObjectName("label_114")
        self.frequency_dm_account_3 = QtWidgets.QSpinBox(self.dm_settings_frame_2)
        self.frequency_dm_account_3.setGeometry(QtCore.QRect(280, 170, 71, 22))
        self.frequency_dm_account_3.setObjectName("frequency_dm_account_3")
        self.pushButton_7 = QtWidgets.QPushButton(self.dm_settings_frame_2)
        self.pushButton_7.setGeometry(QtCore.QRect(280, 210, 75, 23))
        self.pushButton_7.setObjectName("pushButton_7")
        self.warmup_settings_frame = QtWidgets.QFrame(self.tab_3)
        self.warmup_settings_frame.setGeometry(QtCore.QRect(30, 140, 531, 471))
        self.warmup_settings_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.warmup_settings_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.warmup_settings_frame.setObjectName("warmup_settings_frame")
        self.addprofile_button = QtWidgets.QPushButton(self.warmup_settings_frame)
        self.addprofile_button.setGeometry(QtCore.QRect(30, 115, 31, 28))
        self.addprofile_button.setObjectName("addprofile_button")
        self.label_48 = QtWidgets.QLabel(self.warmup_settings_frame)
        self.label_48.setGeometry(QtCore.QRect(30, 180, 171, 20))
        self.label_48.setObjectName("label_48")
        self.label_50 = QtWidgets.QLabel(self.warmup_settings_frame)
        self.label_50.setGeometry(QtCore.QRect(30, 90, 191, 21))
        self.label_50.setObjectName("label_50")
        self.addprofile_checkbox = QtWidgets.QCheckBox(self.warmup_settings_frame)
        self.addprofile_checkbox.setGeometry(QtCore.QRect(30, 70, 171, 17))
        self.addprofile_checkbox.setObjectName("addprofile_checkbox")
        self.label_82 = QtWidgets.QLabel(self.warmup_settings_frame)
        self.label_82.setGeometry(QtCore.QRect(30, 30, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_82.setFont(font)
        self.label_82.setObjectName("label_82")
        self.addbio_textbox = QtWidgets.QTextEdit(self.warmup_settings_frame)
        self.addbio_textbox.setGeometry(QtCore.QRect(30, 210, 471, 51))
        self.addbio_textbox.setObjectName("addbio_textbox")
        self.addprofile_path = QtWidgets.QLabel(self.warmup_settings_frame)
        self.addprofile_path.setGeometry(QtCore.QRect(80, 120, 451, 21))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(7)
        self.addprofile_path.setFont(font)
        self.addprofile_path.setText("")
        self.addprofile_path.setObjectName("addprofile_path")
        self.addbio_checkbox = QtWidgets.QCheckBox(self.warmup_settings_frame)
        self.addbio_checkbox.setGeometry(QtCore.QRect(30, 160, 171, 17))
        self.addbio_checkbox.setObjectName("addbio_checkbox")
        self.addlink_checkbox = QtWidgets.QCheckBox(self.warmup_settings_frame)
        self.addlink_checkbox.setGeometry(QtCore.QRect(30, 280, 151, 17))
        self.addlink_checkbox.setObjectName("addlink_checkbox")
        self.addlink_textbox = QtWidgets.QTextEdit(self.warmup_settings_frame)
        self.addlink_textbox.setGeometry(QtCore.QRect(30, 330, 471, 81))
        self.addlink_textbox.setObjectName("addlink_textbox")
        self.label_85 = QtWidgets.QLabel(self.warmup_settings_frame)
        self.label_85.setGeometry(QtCore.QRect(30, 300, 101, 20))
        self.label_85.setObjectName("label_85")
        self.pushButton = QtWidgets.QPushButton(self.warmup_settings_frame)
        self.pushButton.setGeometry(QtCore.QRect(190, 440, 75, 23))
        self.pushButton.setObjectName("pushButton")
        self.Rough.addTab(self.tab_3, "")
        self.tab_4 = QtWidgets.QWidget()
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        self.tab_4.setFont(font)
        self.tab_4.setObjectName("tab_4")
        self.reel_settings_frame = QtWidgets.QFrame(self.tab_4)
        self.reel_settings_frame.setGeometry(QtCore.QRect(0, 40, 631, 591))
        self.reel_settings_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.reel_settings_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.reel_settings_frame.setObjectName("reel_settings_frame")
        self.taguser_reel_button = QtWidgets.QPushButton(self.reel_settings_frame)
        self.taguser_reel_button.setGeometry(QtCore.QRect(40, 310, 41, 28))
        self.taguser_reel_button.setObjectName("taguser_reel_button")
        self.srcpath_reel_button = QtWidgets.QPushButton(self.reel_settings_frame)
        self.srcpath_reel_button.setGeometry(QtCore.QRect(40, 120, 41, 28))
        self.srcpath_reel_button.setObjectName("srcpath_reel_button")
        self.label_39 = QtWidgets.QLabel(self.reel_settings_frame)
        self.label_39.setGeometry(QtCore.QRect(40, 150, 171, 20))
        self.label_39.setObjectName("label_39")
        self.label_40 = QtWidgets.QLabel(self.reel_settings_frame)
        self.label_40.setGeometry(QtCore.QRect(40, 250, 131, 21))
        self.label_40.setObjectName("label_40")
        self.label_41 = QtWidgets.QLabel(self.reel_settings_frame)
        self.label_41.setGeometry(QtCore.QRect(40, 90, 131, 31))
        self.label_41.setObjectName("label_41")
        self.label_42 = QtWidgets.QLabel(self.reel_settings_frame)
        self.label_42.setGeometry(QtCore.QRect(40, 30, 111, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_42.setFont(font)
        self.label_42.setObjectName("label_42")
        self.caption_reel_textbox = QtWidgets.QTextEdit(self.reel_settings_frame)
        self.caption_reel_textbox.setGeometry(QtCore.QRect(40, 180, 341, 61))
        self.caption_reel_textbox.setObjectName("caption_reel_textbox")
        self.srcpath_reel_label = QtWidgets.QLabel(self.reel_settings_frame)
        self.srcpath_reel_label.setGeometry(QtCore.QRect(90, 125, 521, 21))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(10)
        self.srcpath_reel_label.setFont(font)
        self.srcpath_reel_label.setText("")
        self.srcpath_reel_label.setObjectName("srcpath_reel_label")
        self.location_reel_textbox = QtWidgets.QTextEdit(self.reel_settings_frame)
        self.location_reel_textbox.setGeometry(QtCore.QRect(40, 380, 351, 31))
        self.location_reel_textbox.setObjectName("location_reel_textbox")
        self.label_66 = QtWidgets.QLabel(self.reel_settings_frame)
        self.label_66.setGeometry(QtCore.QRect(40, 350, 71, 21))
        self.label_66.setObjectName("label_66")
        self.taguser_reel_path = QtWidgets.QLabel(self.reel_settings_frame)
        self.taguser_reel_path.setGeometry(QtCore.QRect(90, 310, 521, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(10)
        self.taguser_reel_path.setFont(font)
        self.taguser_reel_path.setText("")
        self.taguser_reel_path.setObjectName("taguser_reel_path")
        self.reeltofeed_checkbox_2 = QtWidgets.QCheckBox(self.reel_settings_frame)
        self.reeltofeed_checkbox_2.setGeometry(QtCore.QRect(150, 250, 121, 21))
        self.reeltofeed_checkbox_2.setObjectName("reeltofeed_checkbox_2")
        self.reeltofeed_checkbox_6 = QtWidgets.QCheckBox(self.reel_settings_frame)
        self.reeltofeed_checkbox_6.setGeometry(QtCore.QRect(40, 70, 121, 21))
        self.reeltofeed_checkbox_6.setObjectName("reeltofeed_checkbox_6")
        self.label_56 = QtWidgets.QLabel(self.reel_settings_frame)
        self.label_56.setGeometry(QtCore.QRect(40, 280, 131, 21))
        self.label_56.setObjectName("label_56")
        self.account_proxy_dial_3 = QtWidgets.QSpinBox(self.reel_settings_frame)
        self.account_proxy_dial_3.setGeometry(QtCore.QRect(150, 280, 91, 22))
        self.account_proxy_dial_3.setObjectName("account_proxy_dial_3")
        self.reeltofeed_checkbox_8 = QtWidgets.QCheckBox(self.reel_settings_frame)
        self.reeltofeed_checkbox_8.setGeometry(QtCore.QRect(130, 350, 121, 21))
        self.reeltofeed_checkbox_8.setObjectName("reeltofeed_checkbox_8")
        self.label_86 = QtWidgets.QLabel(self.reel_settings_frame)
        self.label_86.setGeometry(QtCore.QRect(40, 430, 71, 21))
        self.label_86.setObjectName("label_86")
        self.location_reel_textbox_2 = QtWidgets.QTextEdit(self.reel_settings_frame)
        self.location_reel_textbox_2.setGeometry(QtCore.QRect(130, 430, 111, 31))
        self.location_reel_textbox_2.setObjectName("location_reel_textbox_2")
        self.label_88 = QtWidgets.QLabel(self.reel_settings_frame)
        self.label_88.setGeometry(QtCore.QRect(40, 480, 81, 21))
        self.label_88.setObjectName("label_88")
        self.location_reel_textbox_3 = QtWidgets.QTextEdit(self.reel_settings_frame)
        self.location_reel_textbox_3.setGeometry(QtCore.QRect(130, 480, 111, 31))
        self.location_reel_textbox_3.setObjectName("location_reel_textbox_3")
        self.pushButton_4 = QtWidgets.QPushButton(self.reel_settings_frame)
        self.pushButton_4.setGeometry(QtCore.QRect(310, 530, 75, 23))
        self.pushButton_4.setObjectName("pushButton_4")
        self.post_settings_frame = QtWidgets.QFrame(self.tab_4)
        self.post_settings_frame.setGeometry(QtCore.QRect(650, 40, 551, 651))
        self.post_settings_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.post_settings_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.post_settings_frame.setObjectName("post_settings_frame")
        self.taguserpath_post_button = QtWidgets.QPushButton(self.post_settings_frame)
        self.taguserpath_post_button.setGeometry(QtCore.QRect(40, 320, 41, 28))
        self.taguserpath_post_button.setObjectName("taguserpath_post_button")
        self.srcpath_post_button = QtWidgets.QPushButton(self.post_settings_frame)
        self.srcpath_post_button.setGeometry(QtCore.QRect(40, 120, 41, 28))
        self.srcpath_post_button.setObjectName("srcpath_post_button")
        self.label_51 = QtWidgets.QLabel(self.post_settings_frame)
        self.label_51.setGeometry(QtCore.QRect(40, 150, 171, 20))
        self.label_51.setObjectName("label_51")
        self.label_52 = QtWidgets.QLabel(self.post_settings_frame)
        self.label_52.setGeometry(QtCore.QRect(40, 250, 131, 21))
        self.label_52.setObjectName("label_52")
        self.label_53 = QtWidgets.QLabel(self.post_settings_frame)
        self.label_53.setGeometry(QtCore.QRect(40, 90, 131, 31))
        self.label_53.setObjectName("label_53")
        self.label_54 = QtWidgets.QLabel(self.post_settings_frame)
        self.label_54.setGeometry(QtCore.QRect(40, 30, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_54.setFont(font)
        self.label_54.setObjectName("label_54")
        self.caption_post_textbox = QtWidgets.QTextEdit(self.post_settings_frame)
        self.caption_post_textbox.setGeometry(QtCore.QRect(40, 180, 361, 61))
        self.caption_post_textbox.setObjectName("caption_post_textbox")
        self.srcpath_post_label = QtWidgets.QLabel(self.post_settings_frame)
        self.srcpath_post_label.setGeometry(QtCore.QRect(90, 120, 511, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(10)
        self.srcpath_post_label.setFont(font)
        self.srcpath_post_label.setText("")
        self.srcpath_post_label.setObjectName("srcpath_post_label")
        self.taguserpath_post_label = QtWidgets.QLabel(self.post_settings_frame)
        self.taguserpath_post_label.setGeometry(QtCore.QRect(90, 320, 511, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(10)
        self.taguserpath_post_label.setFont(font)
        self.taguserpath_post_label.setText("")
        self.taguserpath_post_label.setObjectName("taguserpath_post_label")
        self.location_post_textbox = QtWidgets.QTextEdit(self.post_settings_frame)
        self.location_post_textbox.setGeometry(QtCore.QRect(40, 410, 341, 31))
        self.location_post_textbox.setObjectName("location_post_textbox")
        self.label_69 = QtWidgets.QLabel(self.post_settings_frame)
        self.label_69.setGeometry(QtCore.QRect(40, 380, 71, 21))
        self.label_69.setObjectName("label_69")
        self.reeltofeed_checkbox_3 = QtWidgets.QCheckBox(self.post_settings_frame)
        self.reeltofeed_checkbox_3.setGeometry(QtCore.QRect(150, 250, 121, 21))
        self.reeltofeed_checkbox_3.setObjectName("reeltofeed_checkbox_3")
        self.reeltofeed_checkbox_5 = QtWidgets.QCheckBox(self.post_settings_frame)
        self.reeltofeed_checkbox_5.setGeometry(QtCore.QRect(40, 70, 121, 21))
        self.reeltofeed_checkbox_5.setObjectName("reeltofeed_checkbox_5")
        self.reeltofeed_checkbox_7 = QtWidgets.QCheckBox(self.post_settings_frame)
        self.reeltofeed_checkbox_7.setGeometry(QtCore.QRect(270, 70, 121, 21))
        self.reeltofeed_checkbox_7.setObjectName("reeltofeed_checkbox_7")
        self.label_55 = QtWidgets.QLabel(self.post_settings_frame)
        self.label_55.setGeometry(QtCore.QRect(40, 280, 131, 21))
        self.label_55.setObjectName("label_55")
        self.account_proxy_dial_2 = QtWidgets.QSpinBox(self.post_settings_frame)
        self.account_proxy_dial_2.setGeometry(QtCore.QRect(150, 280, 91, 22))
        self.account_proxy_dial_2.setObjectName("account_proxy_dial_2")
        self.reeltofeed_checkbox_9 = QtWidgets.QCheckBox(self.post_settings_frame)
        self.reeltofeed_checkbox_9.setGeometry(QtCore.QRect(110, 380, 121, 21))
        self.reeltofeed_checkbox_9.setObjectName("reeltofeed_checkbox_9")
        self.label_89 = QtWidgets.QLabel(self.post_settings_frame)
        self.label_89.setGeometry(QtCore.QRect(40, 520, 81, 21))
        self.label_89.setObjectName("label_89")
        self.label_90 = QtWidgets.QLabel(self.post_settings_frame)
        self.label_90.setGeometry(QtCore.QRect(40, 470, 71, 21))
        self.label_90.setObjectName("label_90")
        self.location_reel_textbox_4 = QtWidgets.QTextEdit(self.post_settings_frame)
        self.location_reel_textbox_4.setGeometry(QtCore.QRect(130, 470, 111, 31))
        self.location_reel_textbox_4.setObjectName("location_reel_textbox_4")
        self.location_reel_textbox_5 = QtWidgets.QTextEdit(self.post_settings_frame)
        self.location_reel_textbox_5.setGeometry(QtCore.QRect(130, 520, 111, 31))
        self.location_reel_textbox_5.setObjectName("location_reel_textbox_5")
        self.pushButton_5 = QtWidgets.QPushButton(self.post_settings_frame)
        self.pushButton_5.setGeometry(QtCore.QRect(190, 580, 75, 23))
        self.pushButton_5.setObjectName("pushButton_5")
        self.story_settings_frame = QtWidgets.QFrame(self.tab_4)
        self.story_settings_frame.setGeometry(QtCore.QRect(1210, 30, 601, 721))
        self.story_settings_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.story_settings_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.story_settings_frame.setObjectName("story_settings_frame")
        self.taguser_story_button = QtWidgets.QPushButton(self.story_settings_frame)
        self.taguser_story_button.setGeometry(QtCore.QRect(40, 380, 41, 28))
        self.taguser_story_button.setObjectName("taguser_story_button")
        self.srcpath_story_button = QtWidgets.QPushButton(self.story_settings_frame)
        self.srcpath_story_button.setGeometry(QtCore.QRect(40, 120, 41, 28))
        self.srcpath_story_button.setObjectName("srcpath_story_button")
        self.label_70 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_70.setGeometry(QtCore.QRect(40, 150, 171, 20))
        self.label_70.setObjectName("label_70")
        self.label_71 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_71.setGeometry(QtCore.QRect(40, 350, 131, 21))
        self.label_71.setObjectName("label_71")
        self.label_72 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_72.setGeometry(QtCore.QRect(40, 90, 141, 31))
        self.label_72.setObjectName("label_72")
        self.storytohlt_checkbox = QtWidgets.QCheckBox(self.story_settings_frame)
        self.storytohlt_checkbox.setGeometry(QtCore.QRect(230, 520, 201, 21))
        self.storytohlt_checkbox.setObjectName("storytohlt_checkbox")
        self.label_73 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_73.setGeometry(QtCore.QRect(40, 30, 241, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(11)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_73.setFont(font)
        self.label_73.setObjectName("label_73")
        self.text_story_textbox = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox.setGeometry(QtCore.QRect(40, 180, 341, 61))
        self.text_story_textbox.setObjectName("text_story_textbox")
        self.srcpath_story_label = QtWidgets.QLabel(self.story_settings_frame)
        self.srcpath_story_label.setGeometry(QtCore.QRect(90, 120, 501, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(10)
        self.srcpath_story_label.setFont(font)
        self.srcpath_story_label.setText("")
        self.srcpath_story_label.setObjectName("srcpath_story_label")
        self.taguser_story_path = QtWidgets.QLabel(self.story_settings_frame)
        self.taguser_story_path.setGeometry(QtCore.QRect(90, 380, 491, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")
        font.setPointSize(10)
        self.taguser_story_path.setFont(font)
        self.taguser_story_path.setText("")
        self.taguser_story_path.setObjectName("taguser_story_path")
        self.link_story_textbox = QtWidgets.QTextEdit(self.story_settings_frame)
        self.link_story_textbox.setGeometry(QtCore.QRect(40, 440, 341, 61))
        self.link_story_textbox.setObjectName("link_story_textbox")
        self.label_75 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_75.setGeometry(QtCore.QRect(40, 410, 171, 20))
        self.label_75.setObjectName("label_75")
        self.label_77 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_77.setGeometry(QtCore.QRect(40, 520, 121, 16))
        self.label_77.setObjectName("label_77")
        self.label_78 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_78.setGeometry(QtCore.QRect(40, 250, 121, 16))
        self.label_78.setObjectName("label_78")
        self.reeltofeed_checkbox_4 = QtWidgets.QCheckBox(self.story_settings_frame)
        self.reeltofeed_checkbox_4.setGeometry(QtCore.QRect(120, 350, 121, 21))
        self.reeltofeed_checkbox_4.setObjectName("reeltofeed_checkbox_4")
        self.label_74 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_74.setGeometry(QtCore.QRect(40, 270, 31, 21))
        self.label_74.setObjectName("label_74")
        self.label_76 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_76.setGeometry(QtCore.QRect(40, 310, 31, 21))
        self.label_76.setObjectName("label_76")
        self.label_79 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_79.setGeometry(QtCore.QRect(230, 270, 61, 21))
        self.label_79.setObjectName("label_79")
        self.label_80 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_80.setGeometry(QtCore.QRect(230, 310, 61, 21))
        self.label_80.setObjectName("label_80")
        self.label_81 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_81.setGeometry(QtCore.QRect(230, 620, 61, 21))
        self.label_81.setObjectName("label_81")
        self.label_83 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_83.setGeometry(QtCore.QRect(40, 620, 31, 21))
        self.label_83.setObjectName("label_83")
        self.label_84 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_84.setGeometry(QtCore.QRect(40, 560, 31, 21))
        self.label_84.setObjectName("label_84")
        self.label_87 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_87.setGeometry(QtCore.QRect(230, 560, 61, 21))
        self.label_87.setObjectName("label_87")
        self.storytohlt_checkbox_3 = QtWidgets.QCheckBox(self.story_settings_frame)
        self.storytohlt_checkbox_3.setGeometry(QtCore.QRect(140, 410, 201, 21))
        self.storytohlt_checkbox_3.setObjectName("storytohlt_checkbox_3")
        self.storytohlt_checkbox_4 = QtWidgets.QCheckBox(self.story_settings_frame)
        self.storytohlt_checkbox_4.setGeometry(QtCore.QRect(140, 150, 201, 21))
        self.storytohlt_checkbox_4.setObjectName("storytohlt_checkbox_4")
        self.text_story_textbox_2 = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox_2.setGeometry(QtCore.QRect(70, 270, 131, 31))
        self.text_story_textbox_2.setObjectName("text_story_textbox_2")
        self.text_story_textbox_3 = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox_3.setGeometry(QtCore.QRect(70, 310, 131, 31))
        self.text_story_textbox_3.setObjectName("text_story_textbox_3")
        self.text_story_textbox_4 = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox_4.setGeometry(QtCore.QRect(290, 270, 131, 31))
        self.text_story_textbox_4.setObjectName("text_story_textbox_4")
        self.text_story_textbox_5 = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox_5.setGeometry(QtCore.QRect(290, 310, 131, 31))
        self.text_story_textbox_5.setObjectName("text_story_textbox_5")
        self.text_story_textbox_6 = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox_6.setGeometry(QtCore.QRect(60, 560, 131, 31))
        self.text_story_textbox_6.setObjectName("text_story_textbox_6")
        self.text_story_textbox_7 = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox_7.setGeometry(QtCore.QRect(60, 620, 131, 31))
        self.text_story_textbox_7.setObjectName("text_story_textbox_7")
        self.text_story_textbox_8 = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox_8.setGeometry(QtCore.QRect(290, 560, 131, 31))
        self.text_story_textbox_8.setObjectName("text_story_textbox_8")
        self.text_story_textbox_9 = QtWidgets.QTextEdit(self.story_settings_frame)
        self.text_story_textbox_9.setGeometry(QtCore.QRect(290, 620, 131, 31))
        self.text_story_textbox_9.setObjectName("text_story_textbox_9")
        self.pushButton_6 = QtWidgets.QPushButton(self.story_settings_frame)
        self.pushButton_6.setGeometry(QtCore.QRect(210, 680, 75, 23))
        self.pushButton_6.setObjectName("pushButton_6")
        self.account_proxy_dial_5 = QtWidgets.QSpinBox(self.story_settings_frame)
        self.account_proxy_dial_5.setGeometry(QtCore.QRect(330, 350, 91, 22))
        self.account_proxy_dial_5.setObjectName("account_proxy_dial_5")
        self.label_57 = QtWidgets.QLabel(self.story_settings_frame)
        self.label_57.setGeometry(QtCore.QRect(230, 350, 131, 21))
        self.label_57.setObjectName("label_57")
        self.Rough.addTab(self.tab_4, "")
        self.tab_5 = QtWidgets.QWidget()
        self.tab_5.setObjectName("tab_5")
        self.mgmt = QtWidgets.QTabWidget(self.tab_5)
        self.mgmt.setGeometry(QtCore.QRect(0, 90, 1801, 1011))
        self.mgmt.setObjectName("mgmt")
        self.reel_mgmt = QtWidgets.QWidget()
        self.reel_mgmt.setObjectName("reel_mgmt")
        self.label_23 = QtWidgets.QLabel(self.reel_mgmt)
        self.label_23.setGeometry(QtCore.QRect(1375, 0, 87, 31))
        self.label_23.setObjectName("label_23")
        self.totalviews_label = QtWidgets.QLabel(self.reel_mgmt)
        self.totalviews_label.setGeometry(QtCore.QRect(1469, 0, 111, 31))
        self.totalviews_label.setObjectName("totalviews_label")
        self.getreport_reelmgmt_button = QtWidgets.QPushButton(self.reel_mgmt)
        self.getreport_reelmgmt_button.setGeometry(QtCore.QRect(1251, 0, 101, 28))
        self.getreport_reelmgmt_button.setObjectName("getreport_reelmgmt_button")
        self.dtable_2 = QtWidgets.QTableWidget(self.reel_mgmt)
        self.dtable_2.setGeometry(QtCore.QRect(10, 40, 1781, 931))
        self.dtable_2.setAlternatingRowColors(True)
        self.dtable_2.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.dtable_2.setGridStyle(QtCore.Qt.SolidLine)
        self.dtable_2.setObjectName("dtable_2")
        self.dtable_2.setColumnCount(0)
        self.dtable_2.setRowCount(0)
        self.dtable_2.horizontalHeader().setVisible(True)
        self.dtable_2.horizontalHeader().setCascadingSectionResizes(True)
        self.dtable_2.horizontalHeader().setHighlightSections(True)
        self.dtable_2.horizontalHeader().setMinimumSectionSize(37)
        self.dtable_2.horizontalHeader().setSortIndicatorShown(True)
        self.dtable_2.horizontalHeader().setStretchLastSection(True)
        self.dtable_2.verticalHeader().setVisible(True)
        self.dtable_2.verticalHeader().setCascadingSectionResizes(True)
        self.dtable_2.verticalHeader().setSortIndicatorShown(True)
        self.mgmt.addTab(self.reel_mgmt, "")
        self.mentions_mgmt = QtWidgets.QWidget()
        self.mentions_mgmt.setObjectName("mentions_mgmt")
        self.dtable_4 = QtWidgets.QTableWidget(self.mentions_mgmt)
        self.dtable_4.setGeometry(QtCore.QRect(10, 30, 1781, 961))
        self.dtable_4.setAlternatingRowColors(True)
        self.dtable_4.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.dtable_4.setGridStyle(QtCore.Qt.SolidLine)
        self.dtable_4.setObjectName("dtable_4")
        self.dtable_4.setColumnCount(0)
        self.dtable_4.setRowCount(6)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_4.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_4.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_4.setVerticalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_4.setVerticalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_4.setVerticalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_4.setVerticalHeaderItem(5, item)
        self.dtable_4.horizontalHeader().setVisible(True)
        self.dtable_4.horizontalHeader().setCascadingSectionResizes(True)
        self.dtable_4.horizontalHeader().setHighlightSections(True)
        self.dtable_4.horizontalHeader().setMinimumSectionSize(37)
        self.dtable_4.horizontalHeader().setSortIndicatorShown(True)
        self.dtable_4.horizontalHeader().setStretchLastSection(True)
        self.dtable_4.verticalHeader().setVisible(True)
        self.dtable_4.verticalHeader().setCascadingSectionResizes(True)
        self.dtable_4.verticalHeader().setSortIndicatorShown(True)
        self.mgmt.addTab(self.mentions_mgmt, "")
        self.dm_mgmt = QtWidgets.QWidget()
        self.dm_mgmt.setObjectName("dm_mgmt")
        self.dtable_5 = QtWidgets.QTableWidget(self.dm_mgmt)
        self.dtable_5.setGeometry(QtCore.QRect(0, 30, 1791, 961))
        self.dtable_5.setAlternatingRowColors(False)
        self.dtable_5.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.dtable_5.setGridStyle(QtCore.Qt.SolidLine)
        self.dtable_5.setObjectName("dtable_5")
        self.dtable_5.setColumnCount(0)
        self.dtable_5.setRowCount(6)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_5.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_5.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_5.setVerticalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_5.setVerticalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_5.setVerticalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.dtable_5.setVerticalHeaderItem(5, item)
        self.dtable_5.horizontalHeader().setVisible(True)
        self.dtable_5.horizontalHeader().setCascadingSectionResizes(True)
        self.dtable_5.horizontalHeader().setHighlightSections(True)
        self.dtable_5.horizontalHeader().setMinimumSectionSize(37)
        self.dtable_5.horizontalHeader().setSortIndicatorShown(True)
        self.dtable_5.horizontalHeader().setStretchLastSection(True)
        self.dtable_5.verticalHeader().setVisible(True)
        self.dtable_5.verticalHeader().setCascadingSectionResizes(True)
        self.dtable_5.verticalHeader().setSortIndicatorShown(True)
        self.mgmt.addTab(self.dm_mgmt, "")
        self.story_mgmt = QtWidgets.QWidget()
        self.story_mgmt.setObjectName("story_mgmt")
        self.dtable_3 = QtWidgets.QTableWidget(self.story_mgmt)
        self.dtable_3.setGeometry(QtCore.QRect(0, 30, 1791, 951))
        self.dtable_3.setAlternatingRowColors(True)
        self.dtable_3.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.dtable_3.setGridStyle(QtCore.Qt.SolidLine)
        self.dtable_3.setObjectName("dtable_3")
        self.dtable_3.setColumnCount(0)
        self.dtable_3.setRowCount(0)
        self.dtable_3.horizontalHeader().setVisible(True)
        self.dtable_3.horizontalHeader().setCascadingSectionResizes(True)
        self.dtable_3.horizontalHeader().setHighlightSections(True)
        self.dtable_3.horizontalHeader().setMinimumSectionSize(37)
        self.dtable_3.horizontalHeader().setSortIndicatorShown(True)
        self.dtable_3.horizontalHeader().setStretchLastSection(True)
        self.dtable_3.verticalHeader().setVisible(True)
        self.dtable_3.verticalHeader().setCascadingSectionResizes(True)
        self.dtable_3.verticalHeader().setSortIndicatorShown(True)
        self.mgmt.addTab(self.story_mgmt, "")
        self.post_mgmt = QtWidgets.QWidget()
        self.post_mgmt.setObjectName("post_mgmt")
        self.dtable_6 = QtWidgets.QTableWidget(self.post_mgmt)
        self.dtable_6.setGeometry(QtCore.QRect(14, 30, 1771, 961))
        self.dtable_6.setAlternatingRowColors(True)
        self.dtable_6.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.dtable_6.setGridStyle(QtCore.Qt.SolidLine)
        self.dtable_6.setObjectName("dtable_6")
        self.dtable_6.setColumnCount(0)
        self.dtable_6.setRowCount(0)
        self.dtable_6.horizontalHeader().setVisible(True)
        self.dtable_6.horizontalHeader().setCascadingSectionResizes(True)
        self.dtable_6.horizontalHeader().setHighlightSections(True)
        self.dtable_6.horizontalHeader().setMinimumSectionSize(37)
        self.dtable_6.horizontalHeader().setSortIndicatorShown(True)
        self.dtable_6.horizontalHeader().setStretchLastSection(True)
        self.dtable_6.verticalHeader().setVisible(True)
        self.dtable_6.verticalHeader().setCascadingSectionResizes(True)
        self.dtable_6.verticalHeader().setSortIndicatorShown(True)
        self.mgmt.addTab(self.post_mgmt, "")
        self.label_58 = QtWidgets.QLabel(self.tab_5)
        self.label_58.setGeometry(QtCore.QRect(4, 0, 140, 20))
        self.label_58.setObjectName("label_58")
        self.loaddata_mgmt_button = QtWidgets.QPushButton(self.tab_5)
        self.loaddata_mgmt_button.setGeometry(QtCore.QRect(360, 30, 111, 28))
        self.loaddata_mgmt_button.setObjectName("loaddata_mgmt_button")
        self.selected_accmgmt_label = QtWidgets.QLabel(self.tab_5)
        self.selected_accmgmt_label.setGeometry(QtCore.QRect(150, 0, 171, 20))
        self.selected_accmgmt_label.setObjectName("selected_accmgmt_label")
        self.group_acc_combobox_3 = QtWidgets.QComboBox(self.tab_5)
        self.group_acc_combobox_3.setGeometry(QtCore.QRect(140, 30, 161, 26))
        self.group_acc_combobox_3.setObjectName("group_acc_combobox_3")
        self.group_acc_combobox_3.addItem("")
        self.label_24 = QtWidgets.QLabel(self.tab_5)
        self.label_24.setGeometry(QtCore.QRect(10, 30, 115, 21))
        self.label_24.setObjectName("label_24")
        self.Rough.addTab(self.tab_5, "")
        self.verticalLayout.addWidget(self.Rough)
        MWin.setCentralWidget(self.centralWidget)
        self.actionRemove_Accounts = QtWidgets.QAction(MWin)
        self.actionRemove_Accounts.setObjectName("actionRemove_Accounts")
        self.actionWarm_Accounts = QtWidgets.QAction(MWin)
        self.actionWarm_Accounts.setObjectName("actionWarm_Accounts")
        self.actionReels = QtWidgets.QAction(MWin)
        self.actionReels.setObjectName("actionReels")
        self.actionImages = QtWidgets.QAction(MWin)
        self.actionImages.setObjectName("actionImages")
        self.actionStories = QtWidgets.QAction(MWin)
        self.actionStories.setObjectName("actionStories")
        self.actionJO = QtWidgets.QAction(MWin)
        self.actionJO.setObjectName("actionJO")
        self.actionFollow = QtWidgets.QAction(MWin)
        self.actionFollow.setObjectName("actionFollow")
        self.actionLike = QtWidgets.QAction(MWin)
        self.actionLike.setObjectName("actionLike")
        self.actionAdd_New_Accounts_2 = QtWidgets.QAction(MWin)
        self.actionAdd_New_Accounts_2.setObjectName("actionAdd_New_Accounts_2")
        self.actionAdd_New_Proxies = QtWidgets.QAction(MWin)
        self.actionAdd_New_Proxies.setObjectName("actionAdd_New_Proxies")
#-----------------------------------------------------------------------------------------------------------------------------------------
        # ***************************************
        # ACCOUNTS - ACCOUNTS TABLE
        # ***************************
        total_headers = "id email password group_id last_activity status proxy followers following posts profile_pic bio bio_link last_login_time"
        sheet_path_FINAL = 'vps1'
        connection = mysql.connector.connect(host='localhost', user="root", password="Nikzzkr19@",
                                             database=sheet_path_FINAL, auth_plugin='mysql_native_password')
        cursor = connection.cursor()
        query_row_headers = total_headers.split(" ")
        query_temp = ""
        group_id = set()
        group_id_col_number = -1
        for ind, ele in enumerate(query_row_headers):
            if ele == "group_id":
                group_id_col_number = ind
            query_temp += ele + ", "

        #print(query_temp[:-2])
        query = "SELECT {} FROM accounts".format(query_temp[:-2])
        cursor.execute(query)
        global rows
        rows = cursor.fetchall()
        rows = [list(x) for x in rows]
        rows = rows
        col = len(rows[0]) + 1
        row = len(rows) + 1
        # connection.close()
        self.dtable = QtWidgets.QTableWidget(self.tab2)
        self.dtable.setGeometry(QtCore.QRect(0, 70, 1681, 801))
        self.dtable.setAlternatingRowColors(True)
        self.dtable.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.dtable.setGridStyle(QtCore.Qt.SolidLine)
        self.dtable.setObjectName("dtable")

        self.dtable.setColumnCount(int(col))
        self.dtable.setRowCount(int(row))
        row_headers = "checkbox " + total_headers
        for ind, ele in enumerate(row_headers.split()):
            item = QTableWidgetItem()
            item.setText(str(ele))
            self.dtable.setItem(0, ind, item)

        show_col = ""
        for row in rows:
            show_col += str(row[0]) + " "

        for index_r, r in enumerate(rows):
            for index_c, c in enumerate(r):
                item = QTableWidgetItem()
                item.setText(str(c))
                self.dtable.setItem(index_r + 1, index_c + 1, item)
                if index_c == group_id_col_number:
                    group_id.add(c)
        #print(group_id)
        for i in range(self.dtable.rowCount()):
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.Unchecked)
            self.dtable.setItem(i + 1, 0, item)

        self.accounts_allocated_label.setText(str(len(rows)))
        self.acc_selected_label.setText(str(len(rows)))
        self.group_acc_combobox.clear()
        self.changegroup_acc_combobox.clear()
        self.group_acc_combobox.addItem("All")
        self.group_acc_combobox.addItem("None")
        self.changegroup_acc_combobox.addItem("None")
        self.changegroup_acc_combobox.addItem("All")
        self.changegroup_acc_combobox.addItem("Selected")
        self.added_proxy_summary.setText("0")
        self.active_proxy_summary.setText("0")
        self.assigned_proxy_summary.setText("0")
        self.notassigned_proxy_summary.setText("0")

        self.account_proxy_dial_6.setMaximum(999999)
        self.account_proxy_dial_7.setMaximum(999999)
        self.account_proxy_dial_6.setSingleStep(100)
        self.account_proxy_dial_7.setSingleStep(100)
        def clear_table():
            selectall = self.selectall_acc_checkbox.isChecked()
            deselect = self.deselectall_acc_checkbox.isChecked()
            select_by_row = self.selectall_acc_checkbox_2.isChecked()
            # print(selectall, deselect)
            total_headers = "id email password group_id last_activity status proxy followers following posts profile_pic bio bio_link last_login_time"
            connection = mysql.connector.connect(host='localhost', user="root", password="Nikzzkr19@",
                                                 database=sheet_path_FINAL, auth_plugin='mysql_native_password')
            cursor = connection.cursor()
            query_row_headers = total_headers.split(" ")
            query_temp = ""
            group_id = set()
            group_id_col_number = -1
            for ind, ele in enumerate(query_row_headers):
                if ele == "group_id":
                    group_id_col_number = ind
                query_temp += ele + ", "

            query = "SELECT {} FROM accounts".format(query_temp[:-2])
            if select_by_row:
                start_range = self.account_proxy_dial_6.value()
                stop_range = self.account_proxy_dial_7.value()
                query += " where id between {} and {}".format(start_range, stop_range)

            cursor.execute(query)
            global rows
            rows = cursor.fetchall()
            rows = [list(x) for x in rows]
            self.dtable.clear()

            content = self.group_acc_combobox.currentText()
            global temp_rows
            temp_rows = []
            if select_by_row:
                temp_rows = rows
                """start_range = self.account_proxy_dial_6.value()
                stop_range = self.account_proxy_dial_7.value()
                temp_rows = rows[start_range:stop_range + 1]"""
            elif content == "All" or content == "None":
                temp_rows = rows
            else:
                for ele in rows:
                    if ele[group_id_col_number] == content:
                        temp_rows.append(ele)

            self.dtable.setRowCount(int(len(temp_rows)) + 1)
            row_headers = "checkbox " + total_headers

            for ind, ele in enumerate(row_headers.split()):
                item = QTableWidgetItem()
                item.setText(str(ele))
                self.dtable.setItem(0, ind, item)

            show_col = ""
            for row in rows:
                show_col += str(row[0]) + " "
                group_id.add(row[group_id_col_number])

            for index_r, r in enumerate(temp_rows):
                for index_c, c in enumerate(r):
                    item = QTableWidgetItem()
                    item.setText(str(c))
                    self.dtable.setItem(index_r + 1, index_c + 1, item)

            len_temp_rows = len(temp_rows)

            """if select_by_row:
                start_range = self.account_proxy_dial_6.value()
                stop_range = self.account_proxy_dial_7.value()
                for i in range(len(temp_rows[start_range:stop_range])):
                    item = QtWidgets.QTableWidgetItem()
                    item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
                    item.setCheckState(QtCore.Qt.Checked)
                    self.dtable.setItem(i + 1, 0, item)"""

            if content == "None" or deselect:
                for i in range(len(temp_rows)):
                    item = QtWidgets.QTableWidgetItem()
                    item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
                    item.setCheckState(QtCore.Qt.Unchecked)
                    self.dtable.setItem(i + 1, 0, item)

            else:
                for i in range(len(temp_rows)):
                    item = QtWidgets.QTableWidgetItem()
                    item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
                    item.setCheckState(QtCore.Qt.Checked)
                    self.dtable.setItem(i + 1, 0, item)

            self.acc_selected_label.setText(str(len_temp_rows))
            self.group_acc_combobox.clear()
            self.changegroup_acc_combobox.clear()
            self.group_acc_combobox.addItem("All")
            self.group_acc_combobox.addItem("None")
            self.changegroup_acc_combobox.addItem("None")
            self.changegroup_acc_combobox.addItem("All")
            self.changegroup_acc_combobox.addItem("Selected")

            for ele in list(group_id):
                self.group_acc_combobox.addItem(str(ele))
                self.changegroup_acc_combobox.addItem(str(ele))

        def change_group_id():
            try:
                connection = mysql.connector.connect(host='localhost', user="root", password="Nikzzkr19@",
                                                     database=sheet_path_FINAL,
                                                     auth_plugin='mysql_native_password')
                c = connection.cursor()
                current_group = self.changegroup_acc_combobox.currentText()
                target_group = self.groupchange_textedit_textbox_2.toPlainText()
                self.groupchange_textedit_textbox_2.setText("")
                print("Checking for current group : ", current_group)
                global temp_rows
                if current_group == "Selected":
                    items = []
                    for i in range(self.dtable.rowCount()):
                        item = self.dtable.item(i, 0)
                        if item.checkState() == QtCore.Qt.Checked:
                            items.append(item)
                    selected_id = []
                    for it in items:
                        r = it.row()
                        selected_id.append(self.dtable.item(int(r), 1).text())
                    print("Selected ids are : ", selected_id)

                    for ids in selected_id:
                        update_query = """UPDATE accounts SET group_id='{}' WHERE id = {}""".format(target_group, ids)
                        update_story = """UPDATE story SET group_id='{}' WHERE id = {}""".format(target_group, ids)
                        update_reel = """UPDATE reels SET group_id='{}' WHERE id = {}""".format(target_group, ids)
                        update_dm = """UPDATE dm SET group_id='{}' WHERE id = {}""".format(target_group, ids)
                        update_mentions = """UPDATE mentions SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                       ids)
                        update_posts = """UPDATE posts SET group_id='{}' WHERE id = {}""".format(target_group, ids)
                        print(update_query)
                        print(update_story)
                        print(update_reel)
                        print(update_dm)
                        print(update_mentions)
                        print(update_posts)
                        c.execute(update_query)
                        c.execute(update_story)
                        c.execute(update_reel)
                        c.execute(update_dm)
                        c.execute(update_mentions)
                        c.execute(update_posts)
                else:
                    for ele in temp_rows:
                        if current_group == "All":
                            update_query = """UPDATE accounts SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                        ele[0])
                            update_story = """UPDATE story SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                     ele[0])
                            update_reel = """UPDATE reels SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                    ele[0])
                            update_dm = """UPDATE dm SET group_id='{}' WHERE id = {}""".format(target_group, ele[0])
                            update_mentions = """UPDATE mentions SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                           ele[0])
                            update_posts = """UPDATE posts SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                     ele[0])

                        elif ele[group_id_col_number] == current_group:
                            # temp_rows[group_id_col_number] = current_group
                            update_query = """UPDATE accounts SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                        ele[0])
                            update_story = """UPDATE story SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                     ele[0])
                            update_reel = """UPDATE reels SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                    ele[0])
                            update_dm = """UPDATE dm SET group_id='{}' WHERE id = {}""".format(target_group, ele[0])
                            update_mentions = """UPDATE mentions SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                           ele[0])
                            update_posts = """UPDATE posts SET group_id='{}' WHERE id = {}""".format(target_group,
                                                                                                     ele[0])
                        else:
                            continue
                        print(update_query)
                        print(update_story)
                        print(update_reel)
                        print(update_dm)
                        print(update_mentions)
                        print(update_posts)
                        c.execute(update_query)
                        c.execute(update_query)
                        c.execute(update_story)
                        c.execute(update_reel)
                        c.execute(update_dm)
                        c.execute(update_mentions)
                        c.execute(update_posts)
                connection.commit()
            except Exception as e:
                print(e)

        def delete_selected():
            try:
                connection = mysql.connector.connect(host='localhost', user="root", password="Nikzzkr19@",
                                                     database=sheet_path_FINAL,
                                                     auth_plugin='mysql_native_password')
                c = connection.cursor()
                items = []
                for i in range(self.dtable.rowCount()):
                    item = self.dtable.item(i, 0)
                    if item.checkState() == QtCore.Qt.Checked:
                        items.append(item)

                selected_id = []
                selected_acc = []
                for it in items:
                    r = it.row()
                    selected_id.append(self.dtable.item(int(r), 1).text())
                    selected_acc.append(self.dtable.item(int(r), 2).text())
                # print("Selected ids are : ", selected_id)
                display = "Deleting these IDs\n"
                for ind, ele in enumerate(selected_id):
                    display += ele + " " + selected_acc[ind] + "\n"

                global bbb
                bbb = "False"

                def popup_button(i):
                    print(i.text())
                    global bbb
                    if "Yes" in str(i.text()):
                        bbb = "True"

                msg = QMessageBox()
                msg.setWindowTitle("Delete Accounts")
                msg.setText(f"Are you Sure to Delete {len(selected_id)} accounts")
                msg.setIcon(QMessageBox.Question)
                msg.setStandardButtons(QMessageBox.Yes | QMessageBox.Cancel)
                msg.setDefaultButton(QMessageBox.Cancel)
                msg.setDetailedText(display)
                msg.buttonClicked.connect(popup_button)
                x = msg.exec_()
                if bbb == "True":
                    print("Deleting Items")
                else:
                    print("Did not Delete")
                    return

                for ids in selected_id:
                    delete_query = """
                                                                                           DELETE FROM accounts
                                                                                           WHERE id = {}
                                                                                           """.format(ids)
                    print(delete_query)
                    c.execute(delete_query)
                connection.commit()
            except Exception as e:
                print(e)

        def save_settings_tab_1():
            working, notworking, challenge_req, added = 0, 0, 0, 0
            for row in rows:
                if "Working" == row[5]:
                    working += 1
                elif "Not Working" == row[5]:
                    notworking += 1
                elif "Challenge" in row[5]:
                    challenge_req += 1

            self.notallocated_acc_summary.setText("0")

            self.active_acc_summary.setText(str(working))
            self.challenge_acc_summary.setText(str(challenge_req))
            self.added_acc_summary.setText("0")
            self.totalcategory_acc_summary.setText(str(len(group_id)))

            new_group = self.groupchange_textedit_textbox.toPlainText()
            if new_group != "" or new_group != None:
                self.group_acc_combobox_2.addItem(str(new_group))
                self.addgroup_acc_combobox.addItem(str(new_group))

            print("New Group is : ", new_group)

            # self.addgroup_acc_combobox.clear()
            # self.group_acc_combobox_2.clear()
            # for ele in group_id:
            # self.addgroup_acc_combobox.addItem(str(ele))
            # self.group_acc_combobox_2.addItem(str(ele))

        for ele in list(group_id):
            self.group_acc_combobox.addItem(str(ele))
            self.changegroup_acc_combobox.addItem(str(ele))
            self.group_acc_combobox_3.addItem(str(ele))

        self.groupchange_btn.clicked.connect(change_group_id)
        self.data_reload_button.clicked.connect(clear_table)
        self.delete_acc_button.clicked.connect(delete_selected)
        self.groupchange_save_button = QtWidgets.QPushButton(self.account_group_frame)
        self.groupchange_save_button.setGeometry(QtCore.QRect(120, 210, 75, 31))
        self.groupchange_save_button.setText("Save")
        self.groupchange_save_button.setObjectName("groupchange_save_button")
        self.groupchange_save_button.clicked.connect(save_settings_tab_1)
        # $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
        # FIRST LOAD SETTINGS
        working, notworking, challenge_req, added = 0, 0, 0, 0
        for row in rows:
            if "Working" == row[5]:
                working += 1
            elif "Not Working" == row[5]:
                notworking += 1
            elif "Challenge" in row[5]:
                challenge_req += 1
        self.notallocated_acc_summary.setText("0")
        self.active_acc_summary.setText(str(working))
        self.challenge_acc_summary.setText(str(challenge_req))
        self.added_acc_summary.setText("0")
        self.totalcategory_acc_summary.setText(str(len(group_id)))
        for ele in group_id:
            self.addgroup_acc_combobox.addItem(str(ele))
            self.group_acc_combobox_2.addItem(str(ele))
            # self.group_accmgmt_combobox.addItem(str(ele))

        def load_tables():
            try:
                connection = mysql.connector.connect(host='localhost', user="root", password="Nikzzkr19@",
                                                     database=sheet_path_FINAL, auth_plugin='mysql_native_password')
                cursor = connection.cursor()
                current_index = (self.mgmt.currentIndex())
                current_group_id = self.group_acc_combobox_3.currentText()
                # REELS -

                """parent_group_dic = {}
                parent_group_query = "select ID, group_id from accounts where group_id = '{}'".format(current_group_id)
                cursor.execute(parent_group_query)
                qqq = cursor.fetchall()
                if qqq:
                    for ind, ele in enumerate(qqq):
                        parent_group_dic[ele[0]] = ele[1]"""

                total_headers = "id email password group_id status proxy number_of_reels reels_pk reels_views reels_likes reels_link reel_name reels_time"
                table_name = "reels"
                query_row_headers = total_headers.split(" ")
                query_temp = ""
                for ind, ele in enumerate(query_row_headers):
                    query_temp += ele + ", "
                if current_group_id == "All":
                    query = "SELECT {} FROM {}".format(query_temp[:-2], table_name)
                else:
                    query = "SELECT {} FROM {} WHERE group_id = '{}'".format(query_temp[:-2], table_name,
                                                                             current_group_id)
                cursor.execute(query)
                rows_reels = cursor.fetchall()
                self.dtable_2.clear()
                if rows_reels:
                    rows_reels = [list(x) for x in rows_reels]
                    col = len(rows_reels[0])
                    row = len(rows_reels) + 1
                    self.dtable_2.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
                    self.dtable_2.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)
                    self.dtable_2.setAlternatingRowColors(True)
                    self.dtable_2.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
                    self.dtable_2.setColumnCount(int(col))
                    self.dtable_2.setRowCount(int(row))
                    row_headers = total_headers
                    for ind, ele in enumerate(row_headers.split()):
                        item = QTableWidgetItem()
                        item.setText(str(ele))
                        self.dtable_2.setItem(0, ind, item)
                    show_col = ""
                    for row in rows_reels:
                        show_col += str(row[0]) + " "
                    for index_r, r in enumerate(rows_reels):
                        for index_c, c in enumerate(r):
                            item = QTableWidgetItem()
                            item.setText(str(c))
                            self.dtable_2.setItem(index_r + 1, index_c, item)

                # MENTIONS -
                total_headers = "id email password group_id status proxy mentions mention_time"
                table_name = "mentions"
                query_row_headers = total_headers.split(" ")
                query_temp = ""
                for ind, ele in enumerate(query_row_headers):
                    query_temp += ele + ", "
                if current_group_id == "All":
                    query = "SELECT {} FROM {}".format(query_temp[:-2], table_name)
                else:
                    query = "SELECT {} FROM {} WHERE group_id = '{}'".format(query_temp[:-2], table_name,
                                                                             current_group_id)
                cursor.execute(query)
                rows_reels = cursor.fetchall()
                self.dtable_4.clear()
                if rows_reels:
                    rows_reels = [list(x) for x in rows_reels]
                    col = len(rows_reels[0])
                    row = len(rows_reels) + 1
                    self.dtable_4.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
                    self.dtable_4.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)
                    self.dtable_4.setAlternatingRowColors(True)
                    self.dtable_4.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
                    self.dtable_4.setColumnCount(int(col))
                    self.dtable_4.setRowCount(int(row))
                    row_headers = total_headers
                    for ind, ele in enumerate(row_headers.split()):
                        item = QTableWidgetItem()
                        item.setText(str(ele))
                        self.dtable_4.setItem(0, ind, item)
                    show_col = ""
                    for row in rows_reels:
                        show_col += str(row[0]) + " "
                    for index_r, r in enumerate(rows_reels):
                        for index_c, c in enumerate(r):
                            item = QTableWidgetItem()
                            item.setText(str(c))
                            self.dtable_4.setItem(index_r + 1, index_c, item)

                # DM -
                total_headers = "id email password group_id status proxy targetted_users dm_time"
                table_name = "dm"
                query_row_headers = total_headers.split(" ")
                query_temp = ""
                for ind, ele in enumerate(query_row_headers):
                    query_temp += ele + ", "
                if current_group_id == "All":
                    query = "SELECT {} FROM {}".format(query_temp[:-2], table_name)
                else:
                    query = "SELECT {} FROM {} WHERE group_id = '{}'".format(query_temp[:-2], table_name,
                                                                             current_group_id)
                cursor.execute(query)
                rows_reels = cursor.fetchall()
                self.dtable_5.clear()
                if rows_reels:
                    rows_reels = [list(x) for x in rows_reels]
                    col = len(rows_reels[0])
                    row = len(rows_reels) + 1
                    self.dtable_5.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
                    self.dtable_5.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)
                    self.dtable_5.setAlternatingRowColors(True)
                    self.dtable_5.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
                    self.dtable_5.setColumnCount(int(col))
                    self.dtable_5.setRowCount(int(row))
                    row_headers = total_headers
                    for ind, ele in enumerate(row_headers.split()):
                        item = QTableWidgetItem()
                        item.setText(str(ele))
                        self.dtable_5.setItem(0, ind, item)
                    show_col = ""
                    for row in rows_reels:
                        show_col += str(row[0]) + " "
                    for index_r, r in enumerate(rows_reels):
                        for index_c, c in enumerate(r):
                            item = QTableWidgetItem()
                            item.setText(str(c))
                            self.dtable_5.setItem(index_r + 1, index_c, item)

                # STORY -
                total_headers = "id email password group_id status proxy highlights_pk highlights_link story_status funnel_link story_time"
                table_name = "story"
                query_row_headers = total_headers.split(" ")
                query_temp = ""
                for ind, ele in enumerate(query_row_headers):
                    query_temp += ele + ", "
                if current_group_id == "All":
                    query = "SELECT {} FROM {}".format(query_temp[:-2], table_name)
                else:
                    query = "SELECT {} FROM {} WHERE group_id = '{}'".format(query_temp[:-2], table_name,
                                                                             current_group_id)
                cursor.execute(query)
                rows_reels = cursor.fetchall()
                self.dtable_3.clear()
                if rows_reels:
                    rows_reels = [list(x) for x in rows_reels]
                    col = len(rows_reels[0])
                    row = len(rows_reels) + 1
                    self.dtable_3.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
                    self.dtable_3.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)
                    self.dtable_3.setAlternatingRowColors(True)
                    self.dtable_3.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
                    self.dtable_3.setColumnCount(int(col))
                    self.dtable_3.setRowCount(int(row))
                    row_headers = total_headers
                    for ind, ele in enumerate(row_headers.split()):
                        item = QTableWidgetItem()
                        item.setText(str(ele))
                        self.dtable_3.setItem(0, ind, item)
                    show_col = ""
                    for row in rows_reels:
                        show_col += str(row[0]) + " "
                    for index_r, r in enumerate(rows_reels):
                        for index_c, c in enumerate(r):
                            item = QTableWidgetItem()
                            item.setText(str(c))
                            self.dtable_3.setItem(index_r + 1, index_c, item)

                # POST
                total_headers = "id email password group_id status proxy post_pk post_link post_time"
                table_name = "posts"
                query_row_headers = total_headers.split(" ")
                query_temp = ""
                for ind, ele in enumerate(query_row_headers):
                    query_temp += ele + ", "
                if current_group_id == "All":
                    query = "SELECT {} FROM {}".format(query_temp[:-2], table_name)
                else:
                    query = "SELECT {} FROM {} WHERE group_id = '{}'".format(query_temp[:-2], table_name,
                                                                             current_group_id)
                cursor.execute(query)
                rows_reels = cursor.fetchall()
                self.dtable_6.clear()
                if rows_reels:
                    rows_reels = [list(x) for x in rows_reels]
                    col = len(rows_reels[0])
                    row = len(rows_reels) + 1
                    self.dtable_6.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
                    self.dtable_6.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)
                    self.dtable_6.setAlternatingRowColors(True)
                    self.dtable_6.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
                    self.dtable_6.setColumnCount(int(col))
                    self.dtable_6.setRowCount(int(row))
                    row_headers = total_headers
                    for ind, ele in enumerate(row_headers.split()):
                        item = QTableWidgetItem()
                        item.setText(str(ele))
                        self.dtable_6.setItem(0, ind, item)
                    show_col = ""
                    for row in rows_reels:
                        show_col += str(row[0]) + " "
                    for index_r, r in enumerate(rows_reels):
                        for index_c, c in enumerate(r):
                            item = QTableWidgetItem()
                            item.setText(str(c))
                            self.dtable_6.setItem(index_r + 1, index_c, item)
                print("*******************************")
                print("*ACTION COMPLETED SUCCESFULLY!*")
                print("*******************************")
            except Exception as e:
                print(e)

        self.loaddata_mgmt_button.clicked.connect(load_tables)

        # ******************
        # ACTUAL STUFF HERE#
        # ******************
        connection = mysql.connector.connect(host='localhost', user="root", password="Nikzzkr19@",
                                             database=sheet_path_FINAL, auth_plugin='mysql_native_password')
        cursor = connection.cursor()

        try:
            def add_accounts():
                folderpath, filter = QtWidgets.QFileDialog.getOpenFileName(None, 'Open file')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.add_acc_path.setText(folderpath)
                update_settings = f"""UPDATE settings SET accounts = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def add_proxy():
                folderpath, filter = QtWidgets.QFileDialog.getOpenFileName(None, 'Open file')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.add_proxy_path.setText(folderpath)
                update_settings = f"""UPDATE settings SET proxy = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def add_profile_pic_path():
                folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, 'Select Folder')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.addprofile_path.setText(folderpath)
                update_settings = f"""UPDATE settings SET profile_pic = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def add_target_path():
                folderpath, filter = QtWidgets.QFileDialog.getOpenFileName(None, 'Open file')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.targetuser_mention_path.setText(folderpath)
                update_settings = f"""UPDATE settings SET target_mentions = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def add_follow_path():
                folderpath, filter = QtWidgets.QFileDialog.getOpenFileName(None, 'Open file')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.targetuser_dm_path_2.setText(folderpath)
                update_settings = f"""UPDATE settings SET target_follow = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def add_dm_path():
                folderpath, filter = QtWidgets.QFileDialog.getOpenFileName(None, 'Open file')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.targetuser_dm_path.setText(folderpath)
                update_settings = f"""UPDATE settings SET target_dm = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def add_reel():
                folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, 'Select Folder')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.srcpath_reel_label.setText(folderpath)
                update_settings = f"""UPDATE settings SET reel = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def add_post():
                folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, 'Select Folder')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.srcpath_post_label.setText(folderpath)
                update_settings = f"""UPDATE settings SET post = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def add_story():
                folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, 'Select Folder')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.srcpath_story_label.setText(folderpath)
                update_settings = f"""UPDATE settings SET story_pic = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def tag_reel():
                folderpath, filter = QtWidgets.QFileDialog.getOpenFileName(None, 'Open file')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.taguser_reel_path.setText(folderpath)
                update_settings = f"""UPDATE settings SET reel_tag = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def tag_post():
                folderpath, filter = QtWidgets.QFileDialog.getOpenFileName(None, 'Open file')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.taguserpath_post_label.setText(folderpath)
                update_settings = f"""UPDATE settings SET post_tag = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            def tag_story():
                folderpath, filter = QtWidgets.QFileDialog.getOpenFileName(None, 'Open file')
                if folderpath == None: folderpath = 'NO Folder Selected'
                self.taguser_story_path.setText(folderpath)
                update_settings = f"""UPDATE settings SET story_tag = '{folderpath}'"""
                cursor.execute(update_settings)
                connection.commit()

            ###########################
            def warmupx():
                spintax = self.addbio_textbox.toPlainText()
                link_spintax = self.addlink_textbox.toPlainText()
                update_settings = f"""UPDATE settings SET bio_text_spintax = '{spintax}', link_spintax='{link_spintax}'"""
                cursor.execute(update_settings)
                connection.commit()

            def mentionx():
                target_post = self.message_dm_textbox_2.toPlainText()
                mention_per_comment = self.comment_mentions_spinBox.value()
                frequency_mentions = self.frequency_mentions_spinBox.value()
                update_settings = f"""UPDATE settings SET target_post_link_text_file = '{target_post}', mentions_per_comment='{mention_per_comment}', comment_frequency='{frequency_mentions}'"""
                cursor.execute(update_settings)
                connection.commit()

            def dmx():
                message_dm = self.message_dm_textbox.toPlainText()
                frequency_dm = self.frequency_dm_account.value()
                update_settings = f"""UPDATE settings SET dm_spintax = '{message_dm}', dm_frequency='{frequency_dm}'"""
                cursor.execute(update_settings)
                connection.commit()

            def reelx():
                reel_caption = self.caption_reel_textbox.toPlainText()
                reel_number = self.account_proxy_dial_3.value()
                reel_location = self.location_reel_textbox.toPlainText()
                reel_latitude = self.location_reel_textbox_2.toPlainText()
                reel_longitude = self.location_reel_textbox_3.toPlainText()
                update_settings = f"""UPDATE settings SET reel_spintax = '{reel_caption}', reel_number_tag='{reel_number}', reel_location = '{reel_location}', reel_latitude='{reel_latitude}', reel_longitude='{reel_longitude}'"""
                cursor.execute(update_settings)
                connection.commit()

            def postx():
                post_caption = self.caption_post_textbox.toPlainText()
                post_number = self.account_proxy_dial_2.value()
                post_location = self.location_post_textbox.toPlainText()
                post_latitude = self.location_reel_textbox_4.toPlainText()
                post_longitude = self.location_reel_textbox_5.toPlainText()
                update_settings = f"""UPDATE settings SET post_caption = '{post_caption}', post_number_tag='{post_number}', post_location = '{post_location}', post_latitude='{post_latitude}', post_longitude='{post_longitude}'"""
                cursor.execute(update_settings)
                connection.commit()

            def storyx():
                story_spintax = self.text_story_textbox.toPlainText()
                story_spintax_x = self.text_story_textbox_2.toPlainText()
                story_spintax_y = self.text_story_textbox_3.toPlainText()
                story_spintax_width = self.text_story_textbox_4.toPlainText()
                story_spintax_height = self.text_story_textbox_5.toPlainText()
                list_of_links = self.link_story_textbox.toPlainText()
                list_of_links_x = self.text_story_textbox_6.toPlainText()
                list_of_links_y = self.text_story_textbox_7.toPlainText()
                list_of_links_width = self.text_story_textbox_8.toPlainText()
                list_of_links_height = self.text_story_textbox_9.toPlainText()
                update_settings = f"""UPDATE settings SET story_spintax = '{story_spintax}', story_text_x='{story_spintax_x}',
                        story_text_y = '{story_spintax_y}',story_text_width='{story_spintax_width}', story_text_height='{story_spintax_height}',
                        story_list_of_links = '{list_of_links}', story_link_x='{list_of_links_x}',
                        story_link_y = '{list_of_links_y}',story_link_width='{list_of_links_width}', story_link_height='{list_of_links_height}'"""
                cursor.execute(update_settings)
                connection.commit()

            def followx():
                start = self.frequency_dm_account_2.value()
                stop = self.frequency_dm_account_3.value()
                update_settings = f"""UPDATE settings SET start_follow = '{start}', stop_follow='{stop}'"""
                cursor.execute(update_settings)
                connection.commit()

            def accountx():
                use_proxies = self.account_proxy_dial_4.value()
                account_per_proxy = self.account_proxy_dial.value()
                update_settings = f"""UPDATE settings SET proxy_num = '{use_proxies}', accounts_per_proxy='{account_per_proxy}'"""
                cursor.execute(update_settings)
                connection.commit()

            self.add_acc_path_button.clicked.connect(add_accounts)
            self.add_proxy_path_button.clicked.connect(add_proxy)
            self.addprofile_button.clicked.connect(add_profile_pic_path)
            self.targetuser_mention_button.clicked.connect(add_target_path)
            self.targetuser_dm_button_2.clicked.connect(add_follow_path)
            self.targetuser_dm_button.clicked.connect(add_dm_path)

            self.srcpath_reel_button.clicked.connect(add_reel)
            self.taguser_reel_button.clicked.connect(tag_reel)
            self.srcpath_post_button.clicked.connect(add_post)
            self.taguserpath_post_button.clicked.connect(tag_post)
            self.srcpath_story_button.clicked.connect(add_story)
            self.taguser_story_button.clicked.connect(tag_story)

            ########################
            self.pushButton.clicked.connect(warmupx)
            self.pushButton_2.clicked.connect(mentionx)
            self.pushButton_3.clicked.connect(dmx)
            self.pushButton_4.clicked.connect(reelx)
            self.pushButton_5.clicked.connect(postx)
            self.pushButton_6.clicked.connect(storyx)
            self.pushButton_7.clicked.connect(followx)
            self.pushButton_8.clicked.connect(accountx)

            settings_query = """
                                     SELECT * FROM settings
                                     """
            cursor.execute(settings_query)
            settings = cursor.fetchall()
            settings = [list(x) for x in settings]
            settings = settings[0]

            self.add_acc_path.setText(str(settings[0]))
            self.add_proxy_path.setText(str(settings[1]))

            self.addprofile_path.setText(str(settings[2]))
            self.targetuser_mention_path.setText(str(settings[3]))
            self.targetuser_dm_path.setText(str(settings[4]))
            self.targetuser_dm_path_2.setText(str(settings[5]))
            self.srcpath_reel_label.setText(str(settings[6]))
            self.taguser_reel_path.setText(str(settings[7]))

            self.srcpath_post_label.setText(str(settings[8]))
            self.taguserpath_post_label.setText(str(settings[9]))

            self.srcpath_story_label.setText(str(settings[10]))
            self.taguser_story_path.setText(str(settings[11]))

            ########################################################
            self.account_proxy_dial_4.setValue(int(settings[12]))
            self.account_proxy_dial.setValue(int(settings[13]))
            # warmup
            self.addbio_textbox.setText(str(settings[14]))
            self.addlink_textbox.setText(str(settings[15]))
            # mention
            self.message_dm_textbox_2.setText(str(settings[16]))
            self.comment_mentions_spinBox.setValue(int(settings[17]))
            self.frequency_mentions_spinBox.setValue(int(settings[18]))
            # follow
            self.frequency_dm_account_2.setValue(int(settings[19]))
            self.frequency_dm_account_3.setValue(int(settings[20]))
            # dm
            self.message_dm_textbox.setText(str(settings[21]))
            self.frequency_dm_account.setValue(int(settings[22]))
            # reel
            self.caption_reel_textbox.setText(str(settings[23]))
            self.account_proxy_dial_3.setValue(int(settings[24]))
            self.location_reel_textbox.setText(str(settings[25]))
            self.location_reel_textbox_2.setText(str(settings[26]))
            self.location_reel_textbox_3.setText(str(settings[27]))
            # post
            self.caption_post_textbox.setText(str(settings[28]))
            self.account_proxy_dial_2.setValue(int(settings[29]))
            self.location_post_textbox.setText(str(settings[30]))
            self.location_reel_textbox_4.setText(str(settings[41]))
            self.location_reel_textbox_5.setText(str(settings[42]))
            # story
            self.text_story_textbox.setText(str(settings[31]))
            self.text_story_textbox_2.setText(str(settings[32]))
            self.text_story_textbox_3.setText(str(settings[33]))
            self.text_story_textbox_4.setText(str(settings[34]))
            self.text_story_textbox_5.setText(str(settings[35]))
            self.link_story_textbox.setText(str(settings[36]))
            self.text_story_textbox_6.setText(str(settings[37]))
            self.text_story_textbox_7.setText(str(settings[38]))
            self.text_story_textbox_8.setText(str(settings[39]))
            self.text_story_textbox_9.setText(str(settings[40]))

        except Exception as e:
            print(e)

        # TOTAL ACCOUNTS HERE
        proxy_file = self.add_proxy_path.text()
        #proxy_file = open(proxy_file, "r")
        global proxy_code, threaded_proxy_list, proxy_dic

        wb_obj = openpyxl.load_workbook(proxy_file)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        proxy_dic = {}
        for i in range(2, max_row + 1):
            proxy_id = sheet_obj.cell(row=i, column=1)
            proxy_reset = sheet_obj.cell(row=i, column=2)
            proxy_dic[proxy_id.value] = proxy_reset.value

        #threaded_proxy_list = proxy_file.readlines()
        #threaded_proxy_list = [x[:-1] if x[-1] == "\n" else x for x in threaded_proxy_list]
        for i in list(proxy_dic.keys()):
            self.listWidget.addItem(str(i))
        self.active_proxy_summary.setText(str(len(proxy_dic.keys())))
        self.account_proxy_dial_4.setMaximum(int(self.active_proxy_summary.text()))
        print("Proxies used are : ",proxy_dic)

        #

        # threaded_proxy_list = ["Rjenkins313:Audit313@173.92.252.25:2003"]
        # proxy_code = ["3"]

        def book_clicked():
            global proxy_code, threaded_proxy_list, proxy_dic
            use_proxies = self.account_proxy_dial_4.value()
            #threaded_proxy_list = threaded_proxy_list[:use_proxies]
            #proxy_code = proxy_code[:use_proxies + 1]
            proxy_dic = dict(itertools.islice(proxy_dic.items(), 0, int(use_proxies)))

            threaded_proxy_list = proxy_dic.keys()
            proxy_code = proxy_dic

            print(f"using only {use_proxies} number of proxies {threaded_proxy_list} with reset dictionary {proxy_code}")

            f = open('tempfile.txt', 'r+')
            read_from_tmp = f.readlines()
            read_from_tmp = [x[:-1] if x[-1] == "\n" else x for x in read_from_tmp]
            logs_file = open("testing/logs.txt", "a")
            logs_file.write("*****************************************************************************\n")
            for ele in read_from_tmp:
                logs_file.write(ele + "\n")
            logs_file.close()
            items = []
            id_arr = []
            accounts_arr = []
            code = "xae243_default_9980"
            try:
                selected_function = self.activity_combobox.currentText()
                skip_input = False
                if selected_function == "Save Settings" or selected_function == "Initialize" or selected_function == "Reset Proxy":
                    skip_input = True

                if not skip_input:
                    working_test_catcher = input(
                        "Enter 1) To Select Working Accounts \n 2) To Select Non-Working Accounts\n 3) To Select Failed Actions\n")

                number_of_accounts = self.account_proxy_dial.value()  # THE NUMBER OF ACCOUNTS ACTIONS PER PROXY
                print(f"Using {number_of_accounts} Per Proxy")

                for i in range(self.dtable.rowCount()):
                    item = self.dtable.item(i, 0)
                    if item.checkState() == QtCore.Qt.Checked:
                        items.append(item)
                col = self.dtable.columnCount()
                row = self.dtable.rowCount()
                reconstructed = []
                for it in items:
                    r = it.row()
                    c = it.column()
                    temp_row = []
                    for ele in range(1, col):
                        ele = self.dtable.item(int(r), ele).text()
                        temp_row.append(ele)
                    reconstructed.append(temp_row)

                # print(reconstructed)
                frequency = {}
                for ele in reconstructed:
                    error_state = ele[4]
                    working = ele[5]

                    if not skip_input:
                        if working_test_catcher == "2":
                            if working == "Working":
                                continue
                        elif working_test_catcher == "3":
                            if "Error" not in error_state:
                                continue
                        elif working_test_catcher == "1":
                            if working == "Not Working":
                                continue

                    id_arr.append(str(ele[0]))
                    accounts_arr.append(ele[1] + ":" + ele[2])
                    if ele[3] in frequency:
                        frequency[ele[3]] += 1
                    else:
                        frequency[ele[3]] = 1
                temp_key = 0
                for item, key in frequency.items():
                    if temp_key < key:
                        temp_key = key
                        code = item

                print("maximum code is : ", code)
                print("Frequency distribution : ", frequency)

                if code == "xae243_default_9980" and selected_function != "Initialize" and selected_function != "Save Settings" and selected_function != "Reset Proxy":
                    print("No max group Detected from the selected, returning")
                    return

                if selected_function != "Save Settings":
                    f.truncate(0)

                if selected_function == "None":
                    return

                elif selected_function == "Initialize":
                    code = self.addgroup_acc_combobox.currentText()
                    inn = input("Press 1) For Sequential login and 2) For threaded login\n")
                    if inn == "1":
                        threadx = threading.Thread(target=seq_initial_login, args=(
                            sheet_path_FINAL, self.add_acc_path.text(), number_of_accounts, threaded_proxy_list,
                            proxy_code,
                            code))
                        threadx.start()
                    elif inn == "2":
                        threadx = threading.Thread(target=first_login, args=(
                        sheet_path_FINAL, self.add_acc_path.text(), number_of_accounts, threaded_proxy_list, proxy_code,
                        code))
                        threadx.start()

                elif selected_function == "Warmup":
                    profile_pic_checkbox = self.addprofile_checkbox.isChecked()
                    profile_pics_path = self.addprofile_path.text()
                    warmup_spintax = self.addbio_textbox.toPlainText()
                    link_checkbox = self.addlink_checkbox.isChecked()
                    link_spintax = ""
                    if link_checkbox:
                        link_spintax = self.addlink_textbox.toPlainText()
                    pics = []
                    for ele in os.listdir(profile_pics_path):
                        pics.append(ele)
                    # warmup(sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts, threaded_proxy_list,
                    #       proxy_code, pics, profile_pics_path, warmup_spintax, link_checkbox, link_spintax)
                    threadx = threading.Thread(target=warmup, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts, threaded_proxy_list,
                        proxy_code, pics, profile_pics_path, warmup_spintax, link_checkbox, link_spintax, profile_pic_checkbox))
                    threadx.start()
                    # save_excel(sheet_path_FINAL)

                elif selected_function == "Reels":
                    hide_reel = self.reeltofeed_checkbox_6.isChecked()
                    tag_users = self.reeltofeed_checkbox_2.isChecked()
                    enable_location = self.reeltofeed_checkbox_8.isChecked()
                    read_from_mentions = []
                    number_of_tags = 0
                    location_reel = {}
                    if enable_location:
                        location = self.location_reel_textbox.toPlainText()
                        latitude = self.location_reel_textbox_2.toPlainText()
                        longitude = self.location_reel_textbox_3.toPlainText()
                        location_reel["location"] = location
                        location_reel["latitude"] = latitude
                        location_reel["longitude"] = longitude
                    if tag_users:
                        number_of_tags = self.account_proxy_dial_3.value()
                        mention_file = self.taguser_reel_path.text()
                        mention_file = open(mention_file, "r")
                        read_from_mentions = mention_file.readlines()
                        read_from_mentions = [x[:-1] if x[-1] == "\n" else x for x in read_from_mentions]
                    reel_folder_path = self.srcpath_reel_label.text()
                    spintax_reels = self.caption_reel_textbox.toPlainText()
                    # reel_posting(sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts, threaded_proxy_list,
                    #             proxy_code, reel_folder_path, spintax_reels, hide_reel, location_reel, tag_users,
                    #             read_from_mentions, number_of_tags, enable_location)
                    threadx = threading.Thread(target=reel_posting, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts, threaded_proxy_list,
                        proxy_code, reel_folder_path, spintax_reels, hide_reel, location_reel, tag_users,
                        read_from_mentions, number_of_tags, enable_location))
                    threadx.start()
                    # save_excel(sheet_path_FINAL)

                elif selected_function == "Mentions":
                    mention_file = self.targetuser_mention_path.text()
                    mention_file = open(mention_file, "r")
                    read_from_mentions = mention_file.readlines()
                    read_from_mentions = [x[:-1] if x[-1] == "\n" else x for x in read_from_mentions]
                    mentions_post_url = self.message_dm_textbox_2.toPlainText()
                    # mentions(sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts, threaded_proxy_list,
                    #         proxy_code, read_from_mentions, mentions_post_url)
                    threadx = threading.Thread(target=mentions, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts, threaded_proxy_list,
                        proxy_code, read_from_mentions, mentions_post_url))
                    threadx.start()
                    # save_excel(sheet_path_FINAL)

                elif selected_function == "DMs":
                    dm_file = self.targetuser_dm_path.text()
                    dm_file = open(dm_file, "r")
                    read_from_dm_file = dm_file.readlines()
                    read_from_dm_file = [x[:-1] if x[-1] == "\n" else x for x in read_from_dm_file]
                    dm_message = self.message_dm_textbox.toPlainText()
                    # dm(sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts, threaded_proxy_list,
                    #   proxy_code, read_from_dm_file, dm_message)
                    threadx = threading.Thread(target=dm, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts, threaded_proxy_list,
                        proxy_code, read_from_dm_file, dm_message))
                    threadx.start()
                    # save_excel(sheet_path_FINAL)

                elif selected_function == "Stories":
                    media_path = self.srcpath_story_label.text()
                    list_of_links = self.link_story_textbox.toPlainText()
                    list_of_links = list_of_links.replace("\n", " ")
                    list_of_links = list_of_links.split(" ")
                    #print("list of links are ", list_of_links)
                    story_images = []
                    add_to_highlight = self.storytohlt_checkbox.isChecked()
                    story_text_spintax = self.text_story_textbox.toPlainText()
                    enable_text_spintax = self.storytohlt_checkbox_4.isChecked()
                    enable_list_of_links = self.storytohlt_checkbox_3.isChecked()
                    story_spintax_dict, list_of_links_dict = {}, {}
                    story_spintax_dict["enable"] = False
                    list_of_links_dict["enable"] = False
                    story_spintax_dict["highlight"] = add_to_highlight
                    if enable_text_spintax:
                        x = self.text_story_textbox_2.toPlainText()
                        y = self.text_story_textbox_3.toPlainText()
                        width = self.text_story_textbox_4.toPlainText()
                        height = self.text_story_textbox_5.toPlainText()
                        story_spintax_dict["x"] = float(x)
                        story_spintax_dict["y"] = float(y)
                        story_spintax_dict["width"] = float(width)
                        story_spintax_dict["height"] = float(height)
                        story_spintax_dict["enable"] = True
                    if enable_list_of_links:
                        x = self.text_story_textbox_6.toPlainText()
                        y = self.text_story_textbox_7.toPlainText()
                        width = self.text_story_textbox_8.toPlainText()
                        height = self.text_story_textbox_9.toPlainText()
                        list_of_links_dict["x"] = float(x)
                        list_of_links_dict["y"] = float(y)
                        list_of_links_dict["width"] = float(width)
                        list_of_links_dict["height"] = float(height)
                        list_of_links_dict["enable"] = True
                    read_from_mentions = []
                    tag_users = self.reeltofeed_checkbox_4.isChecked()
                    number_of_tags = self.account_proxy_dial_5.value()
                    if tag_users:
                        mention_file = self.taguser_story_path.text()
                        mention_file = open(mention_file, "r")
                        read_from_mentions = mention_file.readlines()
                        read_from_mentions = [x[:-1] if x[-1] == "\n" else x for x in read_from_mentions]
                        #print("Tagged users are : ", read_from_mentions)
                    for ele in os.listdir(media_path):
                        story_images.append(media_path + "/" + str(ele))
                   # print("list of links dict are : ", list_of_links_dict)
                    """threaded_story(sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                                   threaded_proxy_list, proxy_code, story_images, list_of_links, story_text_spintax,
                                   story_spintax_dict, list_of_links_dict, tag_users, number_of_tags, read_from_mentions)"""
                    threadx = threading.Thread(target=threaded_story, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                        threaded_proxy_list, proxy_code, story_images, list_of_links, story_text_spintax,
                        story_spintax_dict, list_of_links_dict, tag_users, number_of_tags, read_from_mentions))
                    threadx.start()
                    # save_excel(sheet_path_FINAL)

                elif selected_function == "Delete Highlights":
                    threadx = threading.Thread(target=highlights_delete, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                        threaded_proxy_list, proxy_code))
                    threadx.start()

                elif selected_function == "Follow":
                    follow_path = self.targetuser_dm_path_2.text()
                    start_follow = self.frequency_dm_account_2.value()
                    stop_follow = self.frequency_dm_account_3.value()
                    """threaded_follow_only(sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                                         threaded_proxy_list, proxy_code, follow_path, start_follow, stop_follow)"""
                    threadx = threading.Thread(target=threaded_follow_only, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                        threaded_proxy_list, proxy_code, follow_path, start_follow, stop_follow))
                    threadx.start()
                    # save_excel(sheet_path_FINAL)

                elif selected_function == "Reset Proxy":
                    for ind, ele in enumerate(threaded_proxy_list):
                        reset_proxy(ele, ind, proxy_code)

                elif selected_function == "Relogin":
                    """threaded_relogin(sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                                     threaded_proxy_list, proxy_code)"""
                    threadx = threading.Thread(target=threaded_relogin, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                        threaded_proxy_list, proxy_code))
                    threadx.start()
                    # save_excel(sheet_path_FINAL)

                elif selected_function == "Save Settings":
                    save_excel(sheet_path_FINAL)

                elif selected_function == "Analytics":
                    # analytics(sheet_path_FINAL, self.group_acc_combobox.currentText())
                    eze = analytics(sheet_path_FINAL, self.group_acc_combobox.currentText())
                    self.totalviews_label.setText(str(eze))

                elif selected_function == "Posts":
                    location_reel = {}
                    tag_users = self.reeltofeed_checkbox_3.isChecked()
                    read_from_mentions = []
                    if tag_users:
                        mention_file = self.taguserpath_post_label.text()
                        mention_file = open(mention_file, "r")
                        read_from_mentions = mention_file.readlines()
                        read_from_mentions = [x[:-1] if x[-1] == "\n" else x for x in read_from_mentions]
                    number_of_tags = self.account_proxy_dial_2.value()
                    enable_location = self.reeltofeed_checkbox_9.isChecked()
                    if enable_location:
                        location = self.location_post_textbox.toPlainText()
                        latitude = self.location_reel_textbox_4.toPlainText()
                        longitude = self.location_reel_textbox_5.toPlainText()
                        location_reel["location"] = location
                        location_reel["latitude"] = latitude
                        location_reel["longitude"] = longitude
                    post_folder = self.srcpath_post_label.text()
                    post_or_video = "post"
                    if self.reeltofeed_checkbox_5.isChecked():
                        post_or_video = "post"
                    if self.reeltofeed_checkbox_7.isChecked():
                        post_or_video = "video"
                    print("post_or_video : ", post_or_video)
                    print("Tagging users : ", tag_users)
                    post_caption_spintax = self.caption_post_textbox.toPlainText()
                    """posts(sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                          threaded_proxy_list, proxy_code, post_folder, post_or_video, post_caption_spintax,
                          location_reel, tag_users, read_from_mentions, number_of_tags, enable_location)"""
                    threadx = threading.Thread(target=posts, args=(
                        sheet_path_FINAL, code, id_arr, accounts_arr, number_of_accounts,
                        threaded_proxy_list, proxy_code, post_folder, post_or_video, post_caption_spintax,
                        location_reel, tag_users, read_from_mentions, number_of_tags, enable_location))
                    threadx.start()
                    # save_excel(sheet_path_FINAL)
            except Exception as e:
                print(e)

        self.start_button.clicked.connect(book_clicked)

        def analytics_alone():
            # eze = threading.Thread(target=analytics, args=(sheet_path_FINAL, self.group_acc_combobox_3.currentText()))
            eze = analytics(sheet_path_FINAL, self.group_acc_combobox_3.currentText())
            self.totalviews_label.setText(str(eze))

        self.getreport_reelmgmt_button.clicked.connect(analytics_alone)

        self.retranslateUi(MWin)
        self.Rough.setCurrentIndex(4)
        self.mgmt.setCurrentIndex(4)
        QtCore.QMetaObject.connectSlotsByName(MWin)

    def retranslateUi(self, MWin):
        _translate = QtCore.QCoreApplication.translate
        MWin.setWindowTitle(_translate("MWin", "Instagram Account Management"))
        self.label_6.setText(_translate("MWin", "Active Accounts: "))
        self.label_28.setText(_translate("MWin", "Not Allocated:"))
        self.label_5.setText(_translate("MWin", "Added : "))
        self.label_7.setText(_translate("MWin", "Challenge Required : "))
        self.label_32.setText(_translate("MWin", "Account Category:"))
        self.label_3.setText(_translate("MWin", "Accounts Summary"))
        self.add_acc_path_button.setText(_translate("MWin", "+"))
        self.label_22.setText(_translate("MWin", "Account Group:"))
        self.addgroup_acc_combobox.setItemText(0, _translate("MWin", "All"))
        self.label_4.setText(_translate("MWin", "Proxies Summary"))
        self.add_proxy_path_button.setText(_translate("MWin", "+"))
        self.label_9.setText(_translate("MWin", "Added : "))
        self.label_10.setText(_translate("MWin", "Active Proxies : "))
        self.label_8.setText(_translate("MWin", "Assigned : "))
        self.label_11.setText(_translate("MWin", "Not Assigned : "))
        self.add_proxy_path_2.setText(_translate("MWin", "Total Proxy List"))
        self.label_127.setText(_translate("MWin", "Accounts per Proxy:"))
        self.label_128.setText(_translate("MWin", "Use Proxies:"))
        self.pushButton_8.setText(_translate("MWin", "SAVE"))
        self.Rough.setTabText(self.Rough.indexOf(self.tab), _translate("MWin", "Dashboard"))
        self.dtable.setSortingEnabled(False)
        self.label.setText(_translate("MWin", "Acc Selected:"))
        self.acc_selected_label.setText(_translate("MWin", "0"))
        self.start_button.setText(_translate("MWin", "Start"))
        self.delete_acc_button.setText(_translate("MWin", "Remove"))
        self.data_reload_button.setText(_translate("MWin", "Reload"))
        self.group_acc_combobox.setItemText(0, _translate("MWin", "All"))
        self.label_20.setText(_translate("MWin", "Account Group:"))
        self.activity_combobox.setItemText(0, _translate("MWin", "None"))
        self.activity_combobox.setItemText(1, _translate("MWin", "Analytics"))
        self.activity_combobox.setItemText(2, _translate("MWin", "Relogin"))
        self.activity_combobox.setItemText(3, _translate("MWin", "Save Settings"))
        self.activity_combobox.setItemText(4, _translate("MWin", "Initialize"))
        self.activity_combobox.setItemText(5, _translate("MWin", "Warmup"))
        self.activity_combobox.setItemText(6, _translate("MWin", "Account Checker"))
        self.activity_combobox.setItemText(7, _translate("MWin", "Reels"))
        self.activity_combobox.setItemText(8, _translate("MWin", "Reset Proxy"))
        self.activity_combobox.setItemText(9, _translate("MWin", "Mentions"))
        self.activity_combobox.setItemText(10, _translate("MWin", "DMs"))
        self.activity_combobox.setItemText(11, _translate("MWin", "Stories"))
        self.activity_combobox.setItemText(12, _translate("MWin", "Posts"))
        self.activity_combobox.setItemText(13, _translate("MWin", "Follow"))
        self.label_21.setText(_translate("MWin", "Activity:"))
        self.label_34.setText(_translate("MWin", "Change Group:"))
        self.changegroup_acc_combobox.setItemText(0, _translate("MWin", "None"))
        self.changegroup_acc_combobox.setItemText(1, _translate("MWin", "All"))
        self.groupchange_btn.setText(_translate("MWin", "Change"))
        self.accounts_allocated_label.setText(_translate("MWin", "0"))
        self.label_44.setText(_translate("MWin", "Accounts Allocated:"))
        self.selectall_acc_checkbox.setText(_translate("MWin", "Select All Accounts"))
        self.deselectall_acc_checkbox.setText(_translate("MWin", "Deselect All Accounts"))
        self.label_38.setText(_translate("MWin", "Target Group:"))
        self.groupchange_textedit_textbox_2.setHtml(_translate("MWin", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'Microsoft YaHei\'; font-size:11pt; font-weight:400; font-style:normal;\">\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p></body></html>"))
        self.label_129.setText(_translate("MWin", "Start Row"))
        self.label_130.setText(_translate("MWin", "End Row"))
        self.selectall_acc_checkbox_2.setText(_translate("MWin", "Select by row number"))
        self.Rough.setTabText(self.Rough.indexOf(self.tab2), _translate("MWin", "Accounts"))
        self.label_35.setText(_translate("MWin", "Account Groups"))
        self.groupchange_textedit_textbox.setHtml(_translate("MWin", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'Microsoft YaHei\'; font-size:11pt; font-weight:400; font-style:normal;\">\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p></body></html>"))
        self.label_36.setText(_translate("MWin", "New Group:"))
        self.label_37.setText(_translate("MWin", "Accounts Allocated:"))
        self.accounts_allocated_label_2.setText(_translate("MWin", "0"))
        self.targetuser_mention_button.setText(_translate("MWin", "+"))
        self.label_91.setText(_translate("MWin", "Target Users Text File : "))
        self.label_92.setText(_translate("MWin", "Mention Settings"))
        self.label_96.setText(_translate("MWin", "Mentions per Comment:"))
        self.label_98.setText(_translate("MWin", "Comment Frequency Per Account:"))
        self.label_94.setText(_translate("MWin", "Target Post link Text File:"))
        self.message_dm_textbox_2.setHtml(_translate("MWin", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'Microsoft YaHei\'; font-size:11pt; font-weight:400; font-style:normal;\">\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p></body></html>"))
        self.pushButton_2.setText(_translate("MWin", "SAVE"))
        self.targetuser_dm_button.setText(_translate("MWin", "+"))
        self.label_95.setText(_translate("MWin", "Target Users Text File : "))
        self.label_99.setText(_translate("MWin", "DM Settings"))
        self.label_102.setText(_translate("MWin", "DM Frequency Per Account:"))
        self.label_104.setText(_translate("MWin", "Message Text Spintax:"))
        self.message_dm_textbox.setHtml(_translate("MWin", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'Microsoft YaHei\'; font-size:11pt; font-weight:400; font-style:normal;\">\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p></body></html>"))
        self.pushButton_3.setText(_translate("MWin", "SAVE"))
        self.targetuser_dm_button_2.setText(_translate("MWin", "+"))
        self.label_97.setText(_translate("MWin", "Target Users Text File : "))
        self.label_100.setText(_translate("MWin", "Follow Settings"))
        self.label_103.setText(_translate("MWin", "Start Follow "))
        self.label_114.setText(_translate("MWin", "Stop Follow"))
        self.pushButton_7.setText(_translate("MWin", "SAVE"))
        self.addprofile_button.setText(_translate("MWin", "+"))
        self.label_48.setText(_translate("MWin", "Bio Text Spintax : "))
        self.label_50.setText(_translate("MWin", "Profile Photo Source File : "))
        self.addprofile_checkbox.setText(_translate("MWin", "Add Profile Photo"))
        self.label_82.setText(_translate("MWin", "Warm-up Settings"))
        self.addbio_checkbox.setText(_translate("MWin", "Add Bio"))
        self.addlink_checkbox.setText(_translate("MWin", "Add External Link"))
        self.label_85.setText(_translate("MWin", "Link Spintax : "))
        self.pushButton.setText(_translate("MWin", "SAVE"))
        self.Rough.setTabText(self.Rough.indexOf(self.tab_3), _translate("MWin", "Settings"))
        self.taguser_reel_button.setText(_translate("MWin", "+"))
        self.srcpath_reel_button.setText(_translate("MWin", "+"))
        self.label_39.setText(_translate("MWin", "Reels Caption Spintax : "))
        self.label_40.setText(_translate("MWin", "Tag Users:"))
        self.label_41.setText(_translate("MWin", "Reels Source File : "))
        self.label_42.setText(_translate("MWin", "Reels Settings"))
        self.label_66.setText(_translate("MWin", "Location:"))
        self.reeltofeed_checkbox_2.setText(_translate("MWin", "Tag Users"))
        self.reeltofeed_checkbox_6.setText(_translate("MWin", "Hide Reel"))
        self.label_56.setText(_translate("MWin", "Number Tag :"))
        self.reeltofeed_checkbox_8.setText(_translate("MWin", "Enable"))
        self.label_86.setText(_translate("MWin", "Latitude:"))
        self.label_88.setText(_translate("MWin", "Longitude:"))
        self.pushButton_4.setText(_translate("MWin", "SAVE"))
        self.taguserpath_post_button.setText(_translate("MWin", "+"))
        self.srcpath_post_button.setText(_translate("MWin", "+"))
        self.label_51.setText(_translate("MWin", "Post Caption Spintax : "))
        self.label_52.setText(_translate("MWin", "Tag Users:"))
        self.label_53.setText(_translate("MWin", "Post Source File : "))
        self.label_54.setText(_translate("MWin", "Post Settings (Photo / Video)"))
        self.label_69.setText(_translate("MWin", "Location:"))
        self.reeltofeed_checkbox_3.setText(_translate("MWin", "Tag Users"))
        self.reeltofeed_checkbox_5.setText(_translate("MWin", "Post"))
        self.reeltofeed_checkbox_7.setText(_translate("MWin", "Video"))
        self.label_55.setText(_translate("MWin", "Number Tag :"))
        self.reeltofeed_checkbox_9.setText(_translate("MWin", "Enable"))
        self.label_89.setText(_translate("MWin", "Longitude:"))
        self.label_90.setText(_translate("MWin", "Latitude:"))
        self.pushButton_5.setText(_translate("MWin", "SAVE"))
        self.taguser_story_button.setText(_translate("MWin", "+"))
        self.srcpath_story_button.setText(_translate("MWin", "+"))
        self.label_70.setText(_translate("MWin", "Text Spintax : "))
        self.label_71.setText(_translate("MWin", "Tag Users:"))
        self.label_72.setText(_translate("MWin", "Story Source File : "))
        self.storytohlt_checkbox.setText(_translate("MWin", "Story to Highlights"))
        self.label_73.setText(_translate("MWin", "Story Settings (Photo / Video)"))
        self.label_75.setText(_translate("MWin", "List of Links"))
        self.label_77.setText(_translate("MWin", "Link Position:"))
        self.label_78.setText(_translate("MWin", "Text Position:"))
        self.reeltofeed_checkbox_4.setText(_translate("MWin", "Tag Users"))
        self.label_74.setText(_translate("MWin", "X:"))
        self.label_76.setText(_translate("MWin", "Y:"))
        self.label_79.setText(_translate("MWin", "Width:"))
        self.label_80.setText(_translate("MWin", "Height:"))
        self.label_81.setText(_translate("MWin", "Height:"))
        self.label_83.setText(_translate("MWin", "Y:"))
        self.label_84.setText(_translate("MWin", "X:"))
        self.label_87.setText(_translate("MWin", "Width:"))
        self.storytohlt_checkbox_3.setText(_translate("MWin", "Add Link"))
        self.storytohlt_checkbox_4.setText(_translate("MWin", "Add Text"))
        self.pushButton_6.setText(_translate("MWin", "SAVE"))
        self.label_57.setText(_translate("MWin", "Number Tag :"))
        self.Rough.setTabText(self.Rough.indexOf(self.tab_4), _translate("MWin", "Upload Settings"))
        self.label_23.setText(_translate("MWin", "Total Views:"))
        self.totalviews_label.setText(_translate("MWin", "0"))
        self.getreport_reelmgmt_button.setText(_translate("MWin", "Get Report"))
        self.dtable_2.setSortingEnabled(False)
        self.mgmt.setTabText(self.mgmt.indexOf(self.reel_mgmt), _translate("MWin", "Reels"))
        self.dtable_4.setSortingEnabled(False)
        item = self.dtable_4.verticalHeaderItem(0)
        item.setText(_translate("MWin", "1"))
        item = self.dtable_4.verticalHeaderItem(1)
        item.setText(_translate("MWin", "2"))
        item = self.dtable_4.verticalHeaderItem(2)
        item.setText(_translate("MWin", "3"))
        item = self.dtable_4.verticalHeaderItem(3)
        item.setText(_translate("MWin", "4"))
        item = self.dtable_4.verticalHeaderItem(4)
        item.setText(_translate("MWin", "5"))
        item = self.dtable_4.verticalHeaderItem(5)
        item.setText(_translate("MWin", "6"))
        self.mgmt.setTabText(self.mgmt.indexOf(self.mentions_mgmt), _translate("MWin", "Mentions"))
        self.dtable_5.setSortingEnabled(False)
        item = self.dtable_5.verticalHeaderItem(0)
        item.setText(_translate("MWin", "1"))
        item = self.dtable_5.verticalHeaderItem(1)
        item.setText(_translate("MWin", "2"))
        item = self.dtable_5.verticalHeaderItem(2)
        item.setText(_translate("MWin", "3"))
        item = self.dtable_5.verticalHeaderItem(3)
        item.setText(_translate("MWin", "4"))
        item = self.dtable_5.verticalHeaderItem(4)
        item.setText(_translate("MWin", "5"))
        item = self.dtable_5.verticalHeaderItem(5)
        item.setText(_translate("MWin", "6"))
        self.mgmt.setTabText(self.mgmt.indexOf(self.dm_mgmt), _translate("MWin", "DMs"))
        self.dtable_3.setSortingEnabled(False)
        self.mgmt.setTabText(self.mgmt.indexOf(self.story_mgmt), _translate("MWin", "Stories-Highlights"))
        self.dtable_6.setSortingEnabled(False)
        self.mgmt.setTabText(self.mgmt.indexOf(self.post_mgmt), _translate("MWin", "Posts"))
        self.label_58.setText(_translate("MWin", "Accounts Selected:"))
        self.loaddata_mgmt_button.setText(_translate("MWin", "load Data"))
        self.selected_accmgmt_label.setText(_translate("MWin", "0"))
        self.group_acc_combobox_3.setItemText(0, _translate("MWin", "All"))
        self.label_24.setText(_translate("MWin", "Account Group:"))
        self.Rough.setTabText(self.Rough.indexOf(self.tab_5), _translate("MWin", "Management"))
        self.actionRemove_Accounts.setText(_translate("MWin", "Remove Accounts"))
        self.actionWarm_Accounts.setText(_translate("MWin", "Warm Accounts"))
        self.actionReels.setText(_translate("MWin", "Reels"))
        self.actionImages.setText(_translate("MWin", "Images"))
        self.actionStories.setText(_translate("MWin", "Stories"))
        self.actionJO.setText(_translate("MWin", "JO"))
        self.actionFollow.setText(_translate("MWin", "Follow Accounts"))
        self.actionLike.setText(_translate("MWin", " Like Media"))
        self.actionAdd_New_Accounts_2.setText(_translate("MWin", "Add New Accounts"))
        self.actionAdd_New_Proxies.setText(_translate("MWin", "Add New Proxies"))

if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MWin = QtWidgets.QMainWindow()
    ui = Ui_MWin()
    ui.setupUi(MWin)
    MWin.show()
    sys.exit(app.exec_())

"""
ALTER TABLE settings ADD proxy_num text;

ALTER TABLE settings ADD accounts_per_proxy text;

ALTER TABLE settings ADD bio_text_spintax text;

ALTER TABLE settings ADD link_spintax text;

ALTER TABLE settings ADD target_post_link_text_file text;

ALTER TABLE settings ADD mentions_per_comment text;

ALTER TABLE settings ADD comment_frequency text;

ALTER TABLE settings ADD start_follow text;

ALTER TABLE settings ADD stop_follow text;

ALTER TABLE settings ADD dm_spintax text;

ALTER TABLE settings ADD dm_frequency text;

ALTER TABLE settings ADD reel_spintax text;

ALTER TABLE settings ADD reel_number_tag text;

ALTER TABLE settings ADD reel_location text;

ALTER TABLE settings ADD reel_latitude text;

ALTER TABLE settings ADD reel_longitude text;

ALTER TABLE settings ADD post_caption text;

ALTER TABLE settings ADD post_number_tag text;

ALTER TABLE settings ADD post_location text;

ALTER TABLE settings ADD story_spintax text;

ALTER TABLE settings ADD story_text_x text;

ALTER TABLE settings ADD story_text_y text;

ALTER TABLE settings ADD story_text_width text;

ALTER TABLE settings ADD story_text_height text;

ALTER TABLE settings ADD story_list_of_links text;

ALTER TABLE settings ADD story_link_x text;

ALTER TABLE settings ADD story_link_y text;

ALTER TABLE settings ADD story_link_width text;

ALTER TABLE settings ADD story_link_height text;

ALTER TABLE settings ADD post_latitude text;

ALTER TABLE settings ADD post_longitude text;
"""
